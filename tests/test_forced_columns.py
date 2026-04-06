from __future__ import annotations

from pathlib import Path

import click
from openpyxl import Workbook, load_workbook

from xlcloak.sanitizer import Sanitizer, parse_full_column_specs


class _FailDetector:
    def detect_cell(self, cell, registry, column_header=None):  # noqa: ANN001
        raise AssertionError(f"Detector should not run for {cell.sheet_name}!{cell.col}{cell.row}")


class _RecordingDetector:
    def __init__(self) -> None:
        self.calls: list[tuple[str, int, int]] = []

    def detect_cell(self, cell, registry, column_header=None):  # noqa: ANN001
        self.calls.append((cell.sheet_name, cell.row, cell.col))
        return [], cell.value


def _make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Email"
    ws["A2"] = "John Smith"
    ws["B2"] = "john@example.com"
    wb.save(path)


def test_parse_full_column_specs_valid() -> None:
    resolved = parse_full_column_specs(("Data.B", "Data.AA"), ["Data"])
    assert ("Data", 2) in resolved
    assert ("Data", 27) in resolved


def test_parse_full_column_specs_rejects_bad_format() -> None:
    try:
        parse_full_column_specs(("Data:B",), ["Data"])
    except click.UsageError as exc:
        assert "Expected format: Sheet.Col" in str(exc)
    else:
        raise AssertionError("Expected UsageError for malformed full-column spec")


def test_parse_full_column_specs_rejects_unknown_sheet() -> None:
    try:
        parse_full_column_specs(("Missing.B",), ["Data"])
    except click.UsageError as exc:
        assert "Unknown sheet" in str(exc)
    else:
        raise AssertionError("Expected UsageError for unknown sheet")


def test_columns_only_requires_full_columns(tmp_path: Path) -> None:
    src = tmp_path / "input.xlsx"
    _make_workbook(src)

    sanitizer = Sanitizer(_FailDetector())
    try:
        sanitizer.run(src, output_path=tmp_path / "out.xlsx", bundle_path=tmp_path / "out.xlcloak", columns_only=True)
    except click.UsageError as exc:
        assert "--columns-only requires" in str(exc)
    else:
        raise AssertionError("Expected UsageError when --columns-only is set without forced columns")


def test_columns_only_skips_detector_and_only_tokenizes_forced_column(tmp_path: Path) -> None:
    src = tmp_path / "input.xlsx"
    _make_workbook(src)

    out = tmp_path / "out.xlsx"
    bundle = tmp_path / "out.xlcloak"
    result = Sanitizer(_FailDetector()).run(
        src,
        output_path=out,
        bundle_path=bundle,
        full_columns=("Data.B",),
        columns_only=True,
    )

    wb = load_workbook(result.sanitized_path)
    ws = wb["Data"]
    assert ws["B1"].value == "Email"
    assert str(ws["B2"].value).startswith("CELL_")
    assert ws["A2"].value == "John Smith"


def test_forced_columns_are_excluded_from_detector_flow(tmp_path: Path) -> None:
    src = tmp_path / "input.xlsx"
    _make_workbook(src)

    detector = _RecordingDetector()
    result = Sanitizer(detector).run(
        src,
        output_path=tmp_path / "out.xlsx",
        bundle_path=tmp_path / "out.xlcloak",
        full_columns=("Data.A",),
    )

    wb = load_workbook(result.sanitized_path)
    ws = wb["Data"]
    assert ws["A1"].value == "Name"
    assert str(ws["A2"].value).startswith("CELL_")
    assert ws["B2"].value == "john@example.com"
    assert all(col != 1 for _sheet, _row, col in detector.calls), (
        f"Detector should not run for forced column A, calls were: {detector.calls}"
    )
