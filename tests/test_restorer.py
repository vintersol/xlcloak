"""Unit tests for xlcloak.restorer — reconciliation engine and report rendering."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from xlcloak.bundle import BundleWriter, DEFAULT_PASSWORD


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _write_bundle(
    path: Path,
    forward_map: dict[str, str],
    reverse_map: dict[str, str],
    password: str = DEFAULT_PASSWORD,
    token_count: int | None = None,
    token_occurrences: dict[str, int] | None = None,
) -> None:
    """Write a minimal bundle for testing."""
    BundleWriter(password).write(
        path,
        forward_map=forward_map,
        reverse_map=reverse_map,
        original_filename="test.xlsx",
        sheets_processed=["Sheet1"],
        token_count=len(forward_map) if token_count is None else token_count,
        token_occurrences=token_occurrences,
    )


def _write_xlsx(path: Path, cells: list[tuple[int, int, str]]) -> None:
    """Write a minimal xlsx with text cells at given (row, col, value)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row, col, value in cells:
        ws.cell(row=row, column=col).value = value
    wb.save(str(path))


# ---------------------------------------------------------------------------
# Test data
# ---------------------------------------------------------------------------

FORWARD_MAP = {
    "John Smith": "PERSON_001",
    "jane@example.com": "EMAIL_002@example.com",
    "Bob Jones": "PERSON_003",
}
REVERSE_MAP = {v: k for k, v in FORWARD_MAP.items()}


# ---------------------------------------------------------------------------
# Tests: derive_restore_paths
# ---------------------------------------------------------------------------


def test_derive_restore_paths_default() -> None:
    """Default naming: data_restored.xlsx and data_restore_manifest.txt."""
    from xlcloak.restorer import derive_restore_paths

    restored, manifest = derive_restore_paths(Path("/tmp/data.xlsx"))
    assert restored == Path("/tmp/data_restored.xlsx")
    assert manifest == Path("/tmp/data_restore_manifest.txt")


def test_derive_restore_paths_with_override() -> None:
    """With output_override, the stem is derived from the override path."""
    from xlcloak.restorer import derive_restore_paths

    restored, manifest = derive_restore_paths(
        Path("/tmp/data.xlsx"),
        output_override=Path("/out/custom.xlsx"),
    )
    assert restored == Path("/out/custom_restored.xlsx")
    assert manifest == Path("/out/custom_restore_manifest.txt")


# ---------------------------------------------------------------------------
# Tests: RestoreResult dataclass
# ---------------------------------------------------------------------------


def test_restore_result_dataclass() -> None:
    """RestoreResult has expected fields."""
    from xlcloak.restorer import RestoreResult

    r = RestoreResult(
        restored_path=Path("/tmp/out.xlsx"),
        manifest_path=Path("/tmp/out_manifest.txt"),
        restored_count=5,
        skipped_count=2,
        new_count=3,
        total_cells=10,
        skipped_cells=[{"token": "PERSON_001", "original": "John Smith"}],
        bundle_version="0.1.0",
        password_mode="default",
    )
    assert r.restored_count == 5
    assert r.skipped_count == 2
    assert r.new_count == 3
    assert r.total_cells == 10
    assert len(r.skipped_cells) == 1


# ---------------------------------------------------------------------------
# Tests: Restorer.run — reconciliation logic
# ---------------------------------------------------------------------------


def test_restorer_restores_token_cells(tmp_path: Path) -> None:
    """Cells containing tokens are restored to their originals."""
    from xlcloak.restorer import Restorer

    # Create sanitized xlsx with tokens
    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [
        (1, 1, "PERSON_001"),
        (2, 1, "EMAIL_002@example.com"),
    ])

    # Write bundle
    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, FORWARD_MAP, REVERSE_MAP)

    result = Restorer().run(sanitized, bundle, force=True)

    assert result.restored_count == 2
    assert result.skipped_count == 1  # Bob Jones token missing from file
    assert result.new_count == 0


def test_restorer_skips_ai_modified_tokens(tmp_path: Path) -> None:
    """Tokens that no longer appear in the file are counted as skipped (AI-modified)."""
    from xlcloak.restorer import Restorer

    # File has only 2 of the 3 original tokens; PERSON_003 (Bob Jones) is gone
    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [
        (1, 1, "PERSON_001"),
        (2, 1, "EMAIL_002@example.com"),
        (3, 1, "AI edited this cell"),  # was PERSON_003, now AI text
    ])

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, FORWARD_MAP, REVERSE_MAP)

    result = Restorer().run(sanitized, bundle, force=True)

    assert result.skipped_count == 1  # PERSON_003 missing from file
    assert "PERSON_003" in [s["token"] for s in result.skipped_cells]


def test_restorer_leaves_non_token_cells_as_new(tmp_path: Path) -> None:
    """Cells that never had tokens (new/unchanged cells) are counted as new."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [
        (1, 1, "PERSON_001"),           # token -> restored
        (2, 1, "Some non-PII text"),    # never a token -> new
        (3, 1, "More unchanged text"),  # never a token -> new
    ])

    bundle = tmp_path / "data.xlcloak"
    forward_map = {"John Smith": "PERSON_001"}
    reverse_map = {"PERSON_001": "John Smith"}
    _write_bundle(bundle, forward_map, reverse_map)

    result = Restorer().run(sanitized, bundle, force=True)

    assert result.restored_count == 1
    assert result.skipped_count == 0
    assert result.new_count == 2


def test_restorer_total_cells_count(tmp_path: Path) -> None:
    """total_cells equals the number of text cells walked in the sanitized file."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [
        (1, 1, "PERSON_001"),
        (2, 1, "regular text"),
        (3, 1, "more text"),
    ])

    bundle = tmp_path / "data.xlcloak"
    forward_map = {"John Smith": "PERSON_001"}
    reverse_map = {"PERSON_001": "John Smith"}
    _write_bundle(bundle, forward_map, reverse_map)

    result = Restorer().run(sanitized, bundle, force=True)

    assert result.total_cells == 3


def test_restorer_restore_result_counts(tmp_path: Path) -> None:
    """RestoreResult counts match the expected reconciliation outcome."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [
        (1, 1, "PERSON_001"),               # restored
        (2, 1, "EMAIL_002@example.com"),    # restored
        (3, 1, "AI changed this"),          # PERSON_003 was here, now AI text -> skipped
        (4, 1, "Never tokenized"),          # new
    ])

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, FORWARD_MAP, REVERSE_MAP)

    result = Restorer().run(sanitized, bundle, force=True)

    assert result.restored_count == 2
    assert result.skipped_count == 1
    assert result.new_count == 2  # "AI changed this" + "Never tokenized"
    assert result.total_cells == 4


def test_restorer_skipped_cells_list(tmp_path: Path) -> None:
    """skipped_cells lists token/original pairs for each missing token."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [
        (1, 1, "PERSON_001"),  # restored; PERSON_003 and EMAIL are missing
    ])

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, FORWARD_MAP, REVERSE_MAP)

    result = Restorer().run(sanitized, bundle, force=True)

    assert result.skipped_count == 2
    tokens = {s["token"] for s in result.skipped_cells}
    assert "PERSON_003" in tokens
    assert "EMAIL_002@example.com" in tokens
    # Each skipped cell has 'original' key
    for sc in result.skipped_cells:
        assert "original" in sc
        assert "token" in sc


def test_restorer_detects_missing_occurrence_for_duplicate_token(tmp_path: Path) -> None:
    """If one duplicate token occurrence is edited away, skipped_count should reflect it."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(
        sanitized,
        [
            (1, 1, "PERSON_001"),
            (2, 1, "AI changed this cell"),
        ],
    )

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(
        bundle,
        forward_map={"John Smith": "PERSON_001"},
        reverse_map={"PERSON_001": "John Smith"},
        token_count=1,
        token_occurrences={"PERSON_001": 2},
    )

    result = Restorer().run(sanitized, bundle, force=True)

    assert result.restored_count == 1
    assert result.skipped_count == 1
    assert result.skipped_cells == [
        {"token": "PERSON_001", "original": "John Smith", "count": 1}
    ]


def test_restorer_wrong_password_raises_value_error(tmp_path: Path) -> None:
    """Wrong password raises ValueError with clear message."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [(1, 1, "PERSON_001")])

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, FORWARD_MAP, REVERSE_MAP, password="correct_password")

    with pytest.raises(ValueError, match="Invalid password"):
        Restorer(password="wrong_password").run(sanitized, bundle)


# ---------------------------------------------------------------------------
# Tests: render_report
# ---------------------------------------------------------------------------


def test_render_report_contains_counts() -> None:
    """render_report output contains restored/skipped/unchanged counts."""
    from xlcloak.restorer import RestoreResult, render_report

    result = RestoreResult(
        restored_path=Path("/tmp/out_restored.xlsx"),
        manifest_path=Path("/tmp/out_restore_manifest.txt"),
        restored_count=35,
        skipped_count=3,
        new_count=4,
        total_cells=42,
        skipped_cells=[
            {"token": "PERSON_001", "original": "John Smith"},
            {"token": "EMAIL_003", "original": "john@example.com"},
            {"token": "PHONE_005", "original": "+1-555-0100"},
        ],
        bundle_version="0.1.0",
        password_mode="default",
    )

    report = render_report(result)

    assert "35" in report
    assert "3" in report
    assert "4" in report
    assert "42" in report
    assert "PERSON_001" in report
    assert "John Smith" in report
    assert "xlcloak restore report" in report.lower() or "restore" in report.lower()


def test_render_report_no_skips() -> None:
    """render_report with no skipped cells omits the skipped tokens section."""
    from xlcloak.restorer import RestoreResult, render_report

    result = RestoreResult(
        restored_path=Path("/tmp/out_restored.xlsx"),
        manifest_path=Path("/tmp/out_restore_manifest.txt"),
        restored_count=10,
        skipped_count=0,
        new_count=2,
        total_cells=12,
        skipped_cells=[],
        bundle_version="0.1.0",
        password_mode="default",
    )

    report = render_report(result)

    assert "10" in report
    assert "0" in report


def test_restorer_writes_restored_file(tmp_path: Path) -> None:
    """Restorer.run() produces a restored xlsx file on disk."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [(1, 1, "PERSON_001")])

    bundle = tmp_path / "data.xlcloak"
    forward_map = {"John Smith": "PERSON_001"}
    reverse_map = {"PERSON_001": "John Smith"}
    _write_bundle(bundle, forward_map, reverse_map)

    result = Restorer().run(sanitized, bundle, force=True)

    assert result.restored_path.exists()
    assert result.manifest_path.exists()


def test_restorer_overwrite_protection(tmp_path: Path) -> None:
    """Restorer.run() without force refuses to overwrite existing output."""
    import click

    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [(1, 1, "PERSON_001")])

    bundle = tmp_path / "data.xlcloak"
    forward_map = {"John Smith": "PERSON_001"}
    reverse_map = {"PERSON_001": "John Smith"}
    _write_bundle(bundle, forward_map, reverse_map)

    Restorer().run(sanitized, bundle, force=True)  # first run

    with pytest.raises(click.UsageError, match="--force"):
        Restorer().run(sanitized, bundle)  # second run without force


# ---------------------------------------------------------------------------
# Tests: substring replacement (mixed-content cells)
# ---------------------------------------------------------------------------


def test_restorer_substring_mixed_content_cell(tmp_path: Path) -> None:
    """Cells with tokens embedded in surrounding text are fully restored."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [
        (1, 1, "Contact PERSON_001 at EMAIL_002@example.com for details"),
    ])

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, FORWARD_MAP, REVERSE_MAP)

    result = Restorer().run(sanitized, bundle, force=True)

    # Cell had 2 tokens — it is a single cell, so restored_count = 1 (cell count)
    assert result.restored_count == 1

    # Load the restored file and verify the cell value
    import openpyxl
    wb = openpyxl.load_workbook(str(result.restored_path))
    ws = wb.active
    cell_value = ws.cell(row=1, column=1).value
    assert cell_value == "Contact John Smith at jane@example.com for details", (
        f"Expected full restoration, got: {cell_value!r}"
    )


def test_restorer_substring_single_token(tmp_path: Path) -> None:
    """Cell with only a token is restored to the original (existing behavior preserved)."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [(1, 1, "PERSON_001")])

    bundle = tmp_path / "data.xlcloak"
    forward_map = {"John Smith": "PERSON_001"}
    reverse_map = {"PERSON_001": "John Smith"}
    _write_bundle(bundle, forward_map, reverse_map)

    result = Restorer().run(sanitized, bundle, force=True)
    assert result.restored_count == 1

    import openpyxl
    wb = openpyxl.load_workbook(str(result.restored_path))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "John Smith"


def test_restorer_substring_no_token_unchanged(tmp_path: Path) -> None:
    """Cell with no token text is left unchanged."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [
        (1, 1, "No PII here at all"),
    ])

    bundle = tmp_path / "data.xlcloak"
    forward_map = {"John Smith": "PERSON_001"}
    reverse_map = {"PERSON_001": "John Smith"}
    _write_bundle(bundle, forward_map, reverse_map)

    result = Restorer().run(sanitized, bundle, force=True)
    assert result.restored_count == 0

    import openpyxl
    wb = openpyxl.load_workbook(str(result.restored_path))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "No PII here at all"


def test_restorer_substring_prefix_collision_resolved(tmp_path: Path) -> None:
    """Longer tokens take precedence over shorter prefix tokens."""
    from xlcloak.restorer import Restorer

    import openpyxl

    # PERSON_001 is a prefix of PERSON_0019 — longer must win
    forward_map = {"John Smith": "PERSON_001", "Jane Doe": "PERSON_0019"}
    reverse_map = {v: k for k, v in forward_map.items()}

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [(1, 1, "PERSON_0019")])

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, forward_map, reverse_map)

    result = Restorer().run(sanitized, bundle, force=True)

    wb = openpyxl.load_workbook(str(result.restored_path))
    ws = wb.active
    cell_value = ws.cell(row=1, column=1).value
    assert cell_value == "Jane Doe", (
        f"Expected 'Jane Doe' (longer token wins), got: {cell_value!r}"
    )


def test_restorer_restored_count_counts_cells_not_tokens(tmp_path: Path) -> None:
    """restored_count is number of cells with substitutions, not total substitutions."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "data_sanitized.xlsx"
    _write_xlsx(sanitized, [
        (1, 1, "PERSON_001 and BOB_003"),  # single cell, two tokens
    ])

    forward_map_local = {"John Smith": "PERSON_001", "Bob Jones": "BOB_003"}
    reverse_map_local = {v: k for k, v in forward_map_local.items()}

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, forward_map_local, reverse_map_local)

    result = Restorer().run(sanitized, bundle, force=True)
    # One cell was patched (even though two tokens were in it)
    assert result.restored_count == 1


# ---------------------------------------------------------------------------
# Tests: edited sanitized workbook scenarios
# ---------------------------------------------------------------------------


def test_restorer_handles_row_reorder_and_insertions(tmp_path: Path) -> None:
    """Row inserts/reorders in sanitized files still restore by token text."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "edited_sanitized.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    ws["A2"] = "PERSON_001"
    ws["A3"] = "EMAIL_002@example.com"
    ws["A4"] = "tail"

    # Simulate spreadsheet edits after sanitize:
    # 1) insert a row near tokens
    # 2) reorder the two token rows
    ws.insert_rows(2, amount=1)
    token_a = ws["A3"].value
    token_b = ws["A4"].value
    ws["A3"] = token_b
    ws["A4"] = token_a
    wb.save(str(sanitized))

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, FORWARD_MAP, REVERSE_MAP)

    result = Restorer().run(sanitized, bundle, force=True)
    assert result.restored_count == 2
    assert result.skipped_count == 1

    restored_wb = openpyxl.load_workbook(str(result.restored_path))
    restored_ws = restored_wb["Sheet1"]
    assert restored_ws["A3"].value == "jane@example.com"
    assert restored_ws["A4"].value == "John Smith"


def test_restorer_counts_deleted_token_rows_as_skipped(tmp_path: Path) -> None:
    """Deleting rows with tokens is reported via skipped_count."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "edited_sanitized.xlsx"
    _write_xlsx(
        sanitized,
        [
            (1, 1, "PERSON_001"),
            (2, 1, "EMAIL_002@example.com"),
            (3, 1, "PERSON_003"),
        ],
    )

    wb = openpyxl.load_workbook(str(sanitized))
    ws = wb["Sheet1"]
    ws.delete_rows(3, 1)
    wb.save(str(sanitized))

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, FORWARD_MAP, REVERSE_MAP)

    result = Restorer().run(sanitized, bundle, force=True)
    assert result.restored_count == 2
    assert result.skipped_count == 1
    assert result.skipped_cells == [{"token": "PERSON_003", "original": "Bob Jones", "count": 1}]


def test_restorer_handles_cut_paste_to_new_location(tmp_path: Path) -> None:
    """Moving a token cell (cut/paste) restores at its new location."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "edited_sanitized.xlsx"
    _write_xlsx(sanitized, [(1, 1, "PERSON_001"), (2, 1, "EMAIL_002@example.com")])

    wb = openpyxl.load_workbook(str(sanitized))
    ws = wb["Sheet1"]
    ws["C5"] = ws["A1"].value
    ws["A1"] = "moved elsewhere"
    wb.save(str(sanitized))

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, FORWARD_MAP, REVERSE_MAP)

    result = Restorer().run(sanitized, bundle, force=True)
    assert result.restored_count == 2
    assert result.skipped_count == 1

    restored_wb = openpyxl.load_workbook(str(result.restored_path))
    restored_ws = restored_wb["Sheet1"]
    assert restored_ws["C5"].value == "John Smith"
    assert restored_ws["A1"].value == "moved elsewhere"


def test_restorer_treats_tampered_token_as_missing(tmp_path: Path) -> None:
    """Manually edited token text is not restored and counts as skipped."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "edited_sanitized.xlsx"
    _write_xlsx(sanitized, [(1, 1, "PERSON_00X"), (2, 1, "EMAIL_002@example.com")])

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, FORWARD_MAP, REVERSE_MAP)

    result = Restorer().run(sanitized, bundle, force=True)
    assert result.restored_count == 1
    assert result.skipped_count == 2
    assert any(entry["token"] == "PERSON_001" for entry in result.skipped_cells)


def test_restorer_preserves_surrounding_user_edits_in_mixed_cell(tmp_path: Path) -> None:
    """When users edit surrounding text, embedded tokens still restore in place."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "edited_sanitized.xlsx"
    _write_xlsx(
        sanitized,
        [(1, 1, "URGENT - contact PERSON_001 before 17:00 via EMAIL_002@example.com")],
    )

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, FORWARD_MAP, REVERSE_MAP)

    result = Restorer().run(sanitized, bundle, force=True)
    assert result.restored_count == 1

    restored_wb = openpyxl.load_workbook(str(result.restored_path))
    restored_ws = restored_wb["Sheet1"]
    assert restored_ws["A1"].value == "URGENT - contact John Smith before 17:00 via jane@example.com"


def test_restorer_handles_sheet_rename(tmp_path: Path) -> None:
    """Renaming a sheet in sanitized workbook should not block restoration."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "edited_sanitized.xlsx"
    _write_xlsx(sanitized, [(1, 1, "PERSON_001")])

    wb = openpyxl.load_workbook(str(sanitized))
    ws = wb["Sheet1"]
    ws.title = "Renamed"
    wb.save(str(sanitized))

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, {"John Smith": "PERSON_001"}, {"PERSON_001": "John Smith"})

    result = Restorer().run(sanitized, bundle, force=True)
    assert result.restored_count == 1

    restored_wb = openpyxl.load_workbook(str(result.restored_path))
    assert "Renamed" in restored_wb.sheetnames
    assert restored_wb["Renamed"]["A1"].value == "John Smith"


def test_restorer_keeps_formula_cells_intact_while_restoring_inputs(tmp_path: Path) -> None:
    """Formula cells remain formulas; referenced token cells get restored."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "edited_sanitized.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "PERSON_001"
    ws["B1"] = '=A1&" ok"'
    wb.save(str(sanitized))

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, {"John Smith": "PERSON_001"}, {"PERSON_001": "John Smith"})

    result = Restorer().run(sanitized, bundle, force=True)
    assert result.restored_count == 1

    restored_wb = openpyxl.load_workbook(str(result.restored_path), data_only=False)
    restored_ws = restored_wb["Sheet1"]
    assert restored_ws["A1"].value == "John Smith"
    assert restored_ws["B1"].data_type == "f"
    assert restored_ws["B1"].value == '=A1&" ok"'


def test_restorer_preserves_data_validation_and_comments(tmp_path: Path) -> None:
    """Workbook metadata survives restore copy-then-patch flow."""
    from openpyxl.comments import Comment
    from openpyxl.worksheet.datavalidation import DataValidation
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "edited_sanitized.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "PERSON_001"
    ws["A1"].comment = Comment("keep me", "qa")
    dv = DataValidation(type="list", formula1='"A,B,C"')
    ws.add_data_validation(dv)
    dv.add("B1")
    wb.save(str(sanitized))

    bundle = tmp_path / "data.xlcloak"
    _write_bundle(bundle, {"John Smith": "PERSON_001"}, {"PERSON_001": "John Smith"})

    result = Restorer().run(sanitized, bundle, force=True)
    restored_wb = openpyxl.load_workbook(str(result.restored_path))
    restored_ws = restored_wb["Sheet1"]
    assert restored_ws["A1"].value == "John Smith"
    assert restored_ws["A1"].comment is not None
    assert restored_ws["A1"].comment.text == "keep me"
    assert restored_ws.data_validations is not None
    assert len(restored_ws.data_validations.dataValidation) == 1


def test_restorer_wrong_bundle_does_not_restore_unrelated_tokens(tmp_path: Path) -> None:
    """Using a non-matching bundle should not restore cells incorrectly."""
    from xlcloak.restorer import Restorer

    sanitized = tmp_path / "edited_sanitized.xlsx"
    _write_xlsx(sanitized, [(1, 1, "PERSON_001"), (2, 1, "EMAIL_002@example.com")])

    wrong_forward = {"Alice Doe": "PERSON_777", "alice@example.com": "EMAIL_888@example.com"}
    wrong_reverse = {v: k for k, v in wrong_forward.items()}
    wrong_bundle = tmp_path / "wrong.xlcloak"
    _write_bundle(wrong_bundle, wrong_forward, wrong_reverse)

    result = Restorer().run(sanitized, wrong_bundle, force=True)
    assert result.restored_count == 0
    assert result.skipped_count == 2

    restored_wb = openpyxl.load_workbook(str(result.restored_path))
    restored_ws = restored_wb["Sheet1"]
    assert restored_ws["A1"].value == "PERSON_001"
    assert restored_ws["A2"].value == "EMAIL_002@example.com"
