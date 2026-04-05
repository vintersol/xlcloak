"""Programmatic fixture generator for xlcloak test suite.

Generates three .xlsx files with graduated PII complexity:
  - simple.xlsx: single sheet, basic PII (names, emails, phones, companies)
  - medium.xlsx: 3 sheets, Swedish PII, cross-sheet references, mixed-content cells
  - hard.xlsx: 5 sheets, formulas, comments, merged cells, chart, edge cases

Run directly to regenerate all fixtures:
    uv run python tests/fixtures/generate_fixtures.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Simple fixture
# ---------------------------------------------------------------------------


def generate_simple(output_dir: Path) -> Path:
    """Generate simple.xlsx — single sheet with basic PII data.

    One sheet named "Contacts" with 20 data rows covering names, emails,
    phones, company names, and non-PII notes. Some rows have gaps (None).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Headers
    ws.append(["Name", "Email", "Phone", "Company", "Notes"])

    # Row data: (name, email, phone, company, notes)
    rows = [
        ("John Smith", "john.smith@acme.com", "+1-555-0101", "Acme Corporation", "Q1 review meeting"),
        ("Maria Garcia", "maria.garcia@globalcorp.net", None, "Global Technologies AB", "Follow up on invoice #1234"),
        ("Erik Andersson", "erik@andersson.se", "+46-70-123-4567", None, "Budget approved for 2026"),
        ("Yuki Tanaka", "y.tanaka@example.jp", "+81-3-1234-5678", "Acme Corporation", "Annual performance review"),
        ("Sarah O'Brien", "sarah.obrien@test.org", None, "Global Technologies AB", "Contract renewal pending"),
        (None, "john.smith@acme.com", "+1-555-0101", "Acme Corporation", "Duplicate contact check"),
        ("John Smith", None, None, None, "No contact info"),
        ("Maria Garcia", "maria.garcia@globalcorp.net", None, "Global Technologies AB", None),
        ("Erik Andersson", None, "+46-70-123-4567", None, "Follow up scheduled"),
        ("Yuki Tanaka", "y.tanaka@example.jp", None, None, "Pending response"),
        ("Sarah O'Brien", "sarah.obrien@test.org", None, "Global Technologies AB", "Meeting on Friday"),
        ("John Smith", "john.smith@acme.com", "+1-555-0101", "Acme Corporation", "Onboarding complete"),
        (None, None, None, "Acme Corporation", "Company-only record"),
        ("Maria Garcia", "maria.garcia@globalcorp.net", "+1-555-0101", None, "Phone updated"),
        ("Erik Andersson", "erik@andersson.se", "+46-70-123-4567", "Global Technologies AB", "New project kick-off"),
        ("Yuki Tanaka", None, "+81-3-1234-5678", "Acme Corporation", None),
        ("Sarah O'Brien", "sarah.obrien@test.org", None, None, "Left voicemail"),
        (None, None, "+1-555-0101", None, "Unknown caller"),
        ("John Smith", "john.smith@acme.com", None, "Acme Corporation", "Invoice approved"),
        ("Maria Garcia", None, None, "Global Technologies AB", "Q4 budget discussion"),
    ]

    for row in rows:
        ws.append(list(row))

    out = output_dir / "simple.xlsx"
    wb.save(str(out))
    return out


# ---------------------------------------------------------------------------
# Medium fixture
# ---------------------------------------------------------------------------


def generate_medium(output_dir: Path) -> Path:
    """Generate medium.xlsx — 3 sheets with Swedish PII and mixed-content cells.

    Sheets: Contacts (50 rows), Transactions (30 rows), Summary (10 rows).
    """
    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Contacts
    # ------------------------------------------------------------------
    ws_contacts = wb.active
    ws_contacts.title = "Contacts"

    # Headers (with intentionally empty H and I header cells for edge case)
    ws_contacts.append([
        "Name", "Email", "Phone", "Company", "SSN", "OrgNumber", "Notes", None, None
    ])

    # Same 5 names as simple fixture (cross-fixture consistency)
    base_contacts = [
        ("John Smith", "john.smith@acme.com", "+1-555-0101", "Acme Corporation", None, None,
         "Q1 review meeting", "extra1", "extra2"),
        ("Maria Garcia", "maria.garcia@globalcorp.net", None, "Global Technologies AB", None, None,
         "Follow up on invoice #1234", None, None),
        ("Erik Andersson", "erik@andersson.se", "+46-70-123-4567", "Volvo AB", None, None,
         "Contact Erik Andersson at erik@andersson.se", None, None),
        ("Yuki Tanaka", "y.tanaka@example.jp", "+81-3-1234-5678", "Acme Corporation", None, None,
         "Annual performance review", None, None),
        ("Sarah O'Brien", "sarah.obrien@test.org", None, "Global Technologies AB", None, None,
         "Contract renewal pending", None, None),
    ]

    # Swedish and additional names
    extra_contacts = [
        ("Bjorn Stromberg", "bjorn.stromberg@volvo.se", "+46-31-100-000", "Volvo AB",
         "199001151234", "556677-8901", "Senior engineer", None, None),
        ("Anna-Karin Lindqvist", "anna@spotify.se", "+46-8-555-0200", "Spotify Technology SA",
         "198507302345", None, "Product manager", None, None),
        ("Lars Eriksson", "lars.eriksson@ericsson.com", "+46-8-719-0000", "Ericsson Ltd",
         None, None, "Account manager", None, None),
        ("Emma Johansson", "emma.j@globalcorp.net", None, "Global Technologies AB",
         None, None, "Call +46-70-123-4567 for Volvo AB", None, None),
        ("Carlos Rodriguez", "carlos@acme.com", "+34-91-000-0001", "Acme Corporation",
         None, None, None, None, None),
        ("Ingrid Holm", "ingrid.holm@ericsson.com", "+46-8-719-0001", "Ericsson Ltd",
         None, None, None, None, None),
        ("Mikael Berg", "mikael.berg@volvo.se", None, "Volvo AB",
         None, None, None, None, None),
        ("Annika Karlsson", None, "+46-70-999-1111", "Spotify Technology SA",
         None, None, None, None, None),
        ("Per Nilsson", "per.nilsson@ericsson.com", None, "Ericsson Ltd",
         None, None, None, None, None),
        ("Lena Svensson", "lena@spotify.se", "+46-8-555-0300", "Spotify Technology SA",
         None, None, None, None, None),
    ]

    all_contacts = base_contacts + extra_contacts

    # Fill to 50 rows by cycling
    for i in range(50):
        row = list(all_contacts[i % len(all_contacts)])
        # Vary some values to avoid pure duplicates
        if i >= len(all_contacts):
            row[0] = f"{row[0]} ({i})"
        ws_contacts.append(row)

    # ------------------------------------------------------------------
    # Sheet 2: Transactions
    # ------------------------------------------------------------------
    ws_tx = wb.create_sheet("Transactions")
    ws_tx.append(["Date", "Amount", "Currency", "Vendor", "Contact Person", "Reference"])

    vendors = [
        "Acme Corporation", "Global Technologies AB", "Volvo AB",
        "Spotify Technology SA", "Ericsson Ltd",
    ]
    contacts = [
        "John Smith", "Maria Garcia", "Erik Andersson", "Bjorn Stromberg", "Lars Eriksson",
    ]
    currencies = ["USD", "SEK", "EUR", "JPY"]

    for i in range(30):
        ws_tx.append([
            f"2026-0{(i % 3) + 1}-{(i % 28) + 1:02d}",
            round(1000.0 + i * 237.5, 2),  # float amounts
            currencies[i % len(currencies)],
            vendors[i % len(vendors)],
            contacts[i % len(contacts)],
            f"REF-{1000 + i}",
        ])

    # ------------------------------------------------------------------
    # Sheet 3: Summary
    # ------------------------------------------------------------------
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["Metric", "Value"])
    summary_rows = [
        ("Total Contacts", 50),
        ("Active Vendors", 5),
        ("Q1 Revenue", 125000),
        ("Q1 Transactions", 30),
        ("Primary Contact", "John Smith"),
        ("Report Portal", "https://internal.acme.com/reports/q1"),
        ("Last Updated", "2026-04-03"),
        ("Prepared By", "Maria Garcia"),
        ("Status", "Final"),
        ("Next Review", "2026-07-01"),
    ]
    for row in summary_rows:
        ws_sum.append(list(row))

    out = output_dir / "medium.xlsx"
    wb.save(str(out))
    return out


# ---------------------------------------------------------------------------
# Hard fixture
# ---------------------------------------------------------------------------


def generate_hard(output_dir: Path) -> Path:
    """Generate hard.xlsx — 5 sheets with formulas, comments, charts, merged cells.

    Sheets: Data, Formulas, Formatting, Charts, EdgeCases.
    """
    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Data — PII + comments + merged cells
    # ------------------------------------------------------------------
    ws_data = wb.active
    ws_data.title = "Data"

    # Merged title row
    ws_data.merge_cells("A1:C1")
    ws_data["A1"] = "Employee Directory"

    ws_data.append(["Name", "Email", "Phone", "Department"])

    names = [
        "John Smith", "Maria Garcia", "Erik Andersson", "Yuki Tanaka", "Sarah O'Brien",
        "Bjorn Stromberg", "Anna-Karin Lindqvist", "Lars Eriksson", "Emma Johansson",
        "Carlos Rodriguez", "Ingrid Holm", "Mikael Berg", "Annika Karlsson",
        "Per Nilsson", "Lena Svensson", "David Brown", "Sophie Martin",
        "Hiroshi Yamamoto", "Isabella Rossi", "Olof Gustafsson",
    ]
    emails = [
        "john.smith@acme.com", "maria.garcia@globalcorp.net", "erik@andersson.se",
        "y.tanaka@example.jp", "sarah.obrien@test.org", "bjorn.stromberg@volvo.se",
        "anna@spotify.se", "lars.eriksson@ericsson.com", "emma.j@globalcorp.net",
        "carlos@acme.com", "ingrid.holm@ericsson.com", "mikael.berg@volvo.se",
        None, "per.nilsson@ericsson.com", "lena@spotify.se", "david.brown@acme.com",
        "sophie.martin@globalcorp.net", None, "isabella.rossi@acme.com",
        "olof.g@volvo.se",
    ]
    phones = [
        "+1-555-0101", None, "+46-70-123-4567", "+81-3-1234-5678", None,
        "+46-31-100-000", "+46-8-555-0200", "+46-8-719-0000", None,
        "+34-91-000-0001", "+46-8-719-0001", None, "+46-70-999-1111",
        None, "+46-8-555-0300", "+1-555-0202", None, "+81-3-9999-0000",
        "+39-06-000-0001", "+46-31-100-001",
    ]
    departments = [
        "Engineering", "Finance", "Engineering", "Product", "Legal",
        "Engineering", "Product", "Sales", "Finance", "Engineering",
        "Engineering", "Engineering", "Product", "Engineering", "Product",
        "Finance", "Legal", "Product", "Finance", "Engineering",
    ]

    for i, (name, email, phone, dept) in enumerate(zip(names, emails, phones, departments)):
        row_num = i + 3  # offset: row 1 = merged title, row 2 = headers
        ws_data.append([name, email, phone, dept])

    # Comments on specific cells (row offsets based on append order)
    # Row 3 = first data row (index 0 = John Smith)
    comment1 = Comment("Review with John Smith", "admin")
    ws_data["A3"].comment = comment1

    comment2 = Comment("PII flagged", "system")
    ws_data["B5"].comment = comment2  # Erik Andersson email

    comment3 = Comment("Check this phone number", "reviewer")
    ws_data["C4"].comment = comment3  # Yuki Tanaka phone

    # Merged cells for department grouping
    ws_data.merge_cells("A20:A22")

    # ------------------------------------------------------------------
    # Sheet 2: Formulas
    # ------------------------------------------------------------------
    ws_formulas = wb.create_sheet("Formulas")

    ws_formulas["A1"] = 100
    ws_formulas["A2"] = 200
    ws_formulas["A3"] = "=SUM(A1:A2)"

    ws_formulas["B1"] = "John Smith"
    ws_formulas["B2"] = "=UPPER(B1)"

    ws_formulas["C1"] = "=TODAY()"

    ws_formulas["D1"] = "=IF(A1>50,\"High\",\"Low\")"

    # ------------------------------------------------------------------
    # Sheet 3: Formatting — data validation and conditional formatting
    # ------------------------------------------------------------------
    ws_fmt = wb.create_sheet("Formatting")

    # Data validation dropdown on column A
    dv = DataValidation(
        type="list",
        formula1='"Option A,Option B,Option C"',
        allow_blank=True,
    )
    dv.sqref = "A1:A20"
    ws_fmt.add_data_validation(dv)

    # Populate some cells in column A with valid choices
    for i, val in enumerate(["Option A", "Option B", "Option C", "Option A", "Option B"], 1):
        ws_fmt.cell(row=i, column=1).value = val

    # Column B: some text data for conditional formatting context
    for i in range(1, 11):
        ws_fmt.cell(row=i, column=2).value = f"Value {i}"

    # ------------------------------------------------------------------
    # Sheet 4: Charts — numeric data + bar chart
    # ------------------------------------------------------------------
    ws_charts = wb.create_sheet("Charts")

    chart_data = [
        ("Month", "Revenue"),
        ("Jan", 12500),
        ("Feb", 15300),
        ("Mar", 18200),
        ("Apr", 14100),
        ("May", 19800),
        ("Jun", 22100),
        ("Jul", 17500),
        ("Aug", 20300),
        ("Sep", 23700),
    ]
    for row in chart_data:
        ws_charts.append(list(row))

    # Create a BarChart from the data
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Revenue"
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Month"

    data_ref = Reference(ws_charts, min_col=2, min_row=1, max_row=10)
    cats_ref = Reference(ws_charts, min_col=1, min_row=2, max_row=10)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws_charts.add_chart(chart, "D1")

    # ------------------------------------------------------------------
    # Sheet 5: EdgeCases
    # ------------------------------------------------------------------
    ws_edge = wb.create_sheet("EdgeCases")

    # A1: empty cell (leave as None)
    ws_edge["A1"] = None

    # A2: very long text with embedded email
    long_text = (
        "Lorem ipsum dolor sit amet consectetur adipiscing elit sed do eiusmod "
        "tempor incididunt ut labore et dolore magna aliqua. Please contact "
        "admin@example.com for support. Ut enim ad minim veniam quis nostrud "
        "exercitation ullamco laboris nisi ut aliquip ex ea commodo consequat. "
        "Duis aute irure dolor in reprehenderit in voluptate velit esse cillum "
        "dolore eu fugiat nulla pariatur excepteur sint occaecat cupidatat non "
        "proident sunt in culpa qui officia deserunt mollit anim id est laborum."
    )
    ws_edge["A2"] = long_text

    # A3-A5: Unicode / international names (ASCII approximations per PLAN note)
    ws_edge["A3"] = "Bjorn Stromberg"  # Swedish (ASCII approx: no umlauts)
    ws_edge["A4"] = "Jose Garcia"       # Spanish (ASCII approx: no accent)
    ws_edge["A5"] = "Li Wei"            # Chinese name in Latin script

    # A6: multi-entity cell
    ws_edge["A6"] = "John Smith (john.smith@acme.com, +1-555-0101) works at Acme Corporation"

    # A7: special characters
    ws_edge["A7"] = "O'Malley & Partners / Dept. of R&D"

    # A8: numeric-looking text stored as string
    ws_edge["A8"] = "12345"

    out = output_dir / "hard.xlsx"
    wb.save(str(out))
    return out


# ---------------------------------------------------------------------------
# Main entrypoint
# ---------------------------------------------------------------------------


if __name__ == "__main__":
    output_dir = Path(__file__).parent

    simple_path = generate_simple(output_dir)
    print(f"Generated: {simple_path}")

    medium_path = generate_medium(output_dir)
    print(f"Generated: {medium_path}")

    hard_path = generate_hard(output_dir)
    print(f"Generated: {hard_path}")
