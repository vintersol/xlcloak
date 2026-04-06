from __future__ import annotations

from pathlib import Path

from click.testing import CliRunner
from openpyxl import Workbook, load_workbook

from xlcloak.cli import main


def _make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Email"
    ws["A2"] = "John Smith"
    ws["B2"] = "john@example.com"
    wb.save(path)


def test_cli_columns_only_requires_full_column(tmp_path: Path) -> None:
    src = tmp_path / "input.xlsx"
    _make_workbook(src)

    runner = CliRunner()
    result = runner.invoke(main, ["sanitize", str(src), "--columns-only"])
    assert result.exit_code != 0
    assert "--columns-only requires at least one --full-column/-c" in result.output


def test_cli_full_column_rejects_bad_format(tmp_path: Path) -> None:
    src = tmp_path / "input.xlsx"
    _make_workbook(src)

    runner = CliRunner()
    result = runner.invoke(main, ["sanitize", str(src), "--columns-only", "-c", "Data:B"])
    assert result.exit_code != 0
    assert "Expected format: Sheet.Col" in result.output


def test_cli_short_c_forced_column_columns_only(tmp_path: Path) -> None:
    src = tmp_path / "input.xlsx"
    _make_workbook(src)

    out_base = tmp_path / "result.xlsx"
    runner = CliRunner()
    result = runner.invoke(
        main,
        ["sanitize", str(src), "--output", str(out_base), "--columns-only", "-c", "Data.B", "--force"],
    )

    assert result.exit_code == 0, result.output
    sanitized_path = tmp_path / "result_sanitized.xlsx"
    wb = load_workbook(sanitized_path)
    ws = wb["Data"]
    assert ws["B1"].value == "Email"
    assert str(ws["B2"].value).startswith("CELL_")
    assert ws["A2"].value == "John Smith"
