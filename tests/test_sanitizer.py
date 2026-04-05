"""Integration tests for the Sanitizer orchestrator."""

from __future__ import annotations

import shutil
from pathlib import Path

import click
import pytest
import spacy.util

from xlcloak.detector import PiiDetector
from xlcloak.sanitizer import SanitizeResult, Sanitizer, derive_output_paths

# Skip entire module if spaCy model is not installed
pytestmark = pytest.mark.skipif(
    not spacy.util.is_package("en_core_web_lg"),
    reason="spaCy model en_core_web_lg not installed",
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

FIXTURES_DIR = Path(__file__).parent / "fixtures"
SIMPLE_FIXTURE = FIXTURES_DIR / "simple.xlsx"


@pytest.fixture(scope="module")
def detector() -> PiiDetector:
    """Return a shared PiiDetector (model loaded once per module)."""
    return PiiDetector(score_threshold=0.4)


# ---------------------------------------------------------------------------
# derive_output_paths tests
# ---------------------------------------------------------------------------


def test_derive_output_paths_default() -> None:
    """Default: paths derived from input path stem."""
    sanitized, bundle, manifest = derive_output_paths(Path("/tmp/data.xlsx"))
    assert sanitized == Path("/tmp/data_sanitized.xlsx")
    assert bundle == Path("/tmp/data.xlcloak")
    assert manifest == Path("/tmp/data_manifest.txt")


def test_derive_output_paths_with_override() -> None:
    """Override: paths derived from output override stem."""
    sanitized, bundle, manifest = derive_output_paths(
        input_path=Path("/tmp/data.xlsx"),
        output_override=Path("/out/custom.xlsx"),
    )
    assert sanitized == Path("/out/custom_sanitized.xlsx")
    assert bundle == Path("/out/custom.xlcloak")
    assert manifest == Path("/out/custom_manifest.txt")


# ---------------------------------------------------------------------------
# Sanitizer.run integration tests
# ---------------------------------------------------------------------------


def test_sanitize_produces_three_files(tmp_path: Path, detector: PiiDetector) -> None:
    """Running sanitize on a fixture produces sanitized xlsx, bundle, and manifest."""
    input_path = tmp_path / "simple.xlsx"
    shutil.copy2(SIMPLE_FIXTURE, input_path)

    result = Sanitizer(detector).run(input_path)

    assert isinstance(result, SanitizeResult)
    assert result.sanitized_path.exists(), "Sanitized xlsx not created"
    assert result.bundle_path.exists(), "Bundle file not created"
    assert result.manifest_path.exists(), "Manifest file not created"


def test_sanitize_overwrite_protection(tmp_path: Path, detector: PiiDetector) -> None:
    """Second run without --force raises UsageError."""
    input_path = tmp_path / "simple.xlsx"
    shutil.copy2(SIMPLE_FIXTURE, input_path)

    sanitizer = Sanitizer(detector)
    sanitizer.run(input_path)  # first run

    with pytest.raises(click.UsageError, match="--force"):
        sanitizer.run(input_path)  # second run without force


def test_sanitize_overwrite_with_force(tmp_path: Path, detector: PiiDetector) -> None:
    """Second run with force=True succeeds without error."""
    input_path = tmp_path / "simple.xlsx"
    shutil.copy2(SIMPLE_FIXTURE, input_path)

    sanitizer = Sanitizer(detector)
    sanitizer.run(input_path)  # first run
    result = sanitizer.run(input_path, force=True)  # second run with force

    assert result.sanitized_path.exists()
    assert result.bundle_path.exists()
    assert result.manifest_path.exists()


def test_manifest_written(tmp_path: Path, detector: PiiDetector) -> None:
    """Manifest file contains expected header and entity breakdown section."""
    input_path = tmp_path / "simple.xlsx"
    shutil.copy2(SIMPLE_FIXTURE, input_path)

    result = Sanitizer(detector).run(input_path)

    manifest_text = result.manifest_path.read_text()
    assert "xlcloak Manifest" in manifest_text
    assert "Entity breakdown:" in manifest_text


def test_manifest_counts_match_sanitize_result(tmp_path: Path, detector: PiiDetector) -> None:
    """Manifest counters should match SanitizeResult counters exactly."""
    input_path = tmp_path / "simple.xlsx"
    shutil.copy2(SIMPLE_FIXTURE, input_path)

    result = Sanitizer(detector).run(input_path)

    manifest_lines = result.manifest_path.read_text().splitlines()
    cells_line = next(line for line in manifest_lines if line.startswith("Cells sanitized:"))
    tokens_line = next(line for line in manifest_lines if line.startswith("Tokens generated:"))
    manifest_cells = int(cells_line.split(":", 1)[1].strip())
    manifest_tokens = int(tokens_line.split(":", 1)[1].strip())

    assert manifest_cells == result.cells_sanitized
    assert manifest_tokens == result.token_count


# ---------------------------------------------------------------------------
# hide_all mode tests
# ---------------------------------------------------------------------------


def test_sanitize_hide_all_replaces_all_cells(tmp_path, simple_fixture):
    """hide_all=True must produce a patch for every text cell in the workbook."""
    from xlcloak.excel_io import WorkbookReader

    output = tmp_path / "out.xlsx"
    bundle = tmp_path / "out.xlcloak"
    detector = PiiDetector()
    sanitizer = Sanitizer(detector)
    result = sanitizer.run(simple_fixture, output_path=output, bundle_path=bundle, hide_all=True)

    # Count expected text cells
    reader = WorkbookReader(simple_fixture)
    wb = reader.open()
    expected_count = sum(1 for _ in reader.iter_text_cells(wb))

    assert result.cells_sanitized == expected_count, (
        f"hide_all should replace all {expected_count} cells, got {result.cells_sanitized}"
    )


def test_sanitize_hide_all_uses_stable_tokens(tmp_path, simple_fixture):
    """Same cell value must map to the same CELL_NNNN token in two hide-all runs."""
    from openpyxl import load_workbook

    detector = PiiDetector()

    out1 = tmp_path / "run1_sanitized.xlsx"
    bundle1 = tmp_path / "run1.xlcloak"
    sanitizer1 = Sanitizer(detector)
    result1 = sanitizer1.run(simple_fixture, output_path=out1, bundle_path=bundle1, hide_all=True)

    out2 = tmp_path / "run2_sanitized.xlsx"
    bundle2 = tmp_path / "run2.xlcloak"
    sanitizer2 = Sanitizer(detector)
    result2 = sanitizer2.run(simple_fixture, output_path=out2, bundle_path=bundle2, hide_all=True, force=True)

    # First text cell in both outputs must have the same token
    wb1 = load_workbook(result1.sanitized_path)
    wb2 = load_workbook(result2.sanitized_path)
    ws1 = wb1.active
    ws2 = wb2.active
    assert ws1.cell(1, 1).value == ws2.cell(1, 1).value, (
        "Same input cell must produce same token across runs"
    )


# ---------------------------------------------------------------------------
# Header pre-pass and row-1 skip tests (Phase 4, Plan 03)
# ---------------------------------------------------------------------------


def test_sanitize_row1_is_scanned(tmp_path):
    """Row-1 text cells are scanned and tokenized when they contain PII."""
    from openpyxl import Workbook, load_workbook
    from xlcloak.detector import PiiDetector
    from xlcloak.sanitizer import Sanitizer

    src = tmp_path / "row1.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "John Smith"
    wb.save(src)

    detector = PiiDetector()
    sanitizer = Sanitizer(detector)
    out = tmp_path / "out.xlsx"
    bundle = tmp_path / "out.xlcloak"
    result = sanitizer.run(src, output_path=out, bundle_path=bundle)

    out_wb = load_workbook(result.sanitized_path)
    value = out_wb["Sheet1"]["A1"].value
    assert isinstance(value, str)
    assert value.startswith("PERSON_"), f"Expected row-1 value to be tokenized, got {value!r}"


def test_sanitize_medium_fixture_hide_all_integration(tmp_path):
    """Full integration: medium fixture sanitized with --hide-all must succeed.

    The medium fixture contains synthetic (Luhn-invalid) Swedish PII and company names.
    hide_all=True bypasses detection, so Luhn-invalid values are not an issue.
    This test is the Phase 4 gate.
    """
    from xlcloak.detector import PiiDetector
    from xlcloak.sanitizer import Sanitizer

    medium_fixture = Path(__file__).parent / "fixtures" / "medium.xlsx"
    output = tmp_path / "medium_sanitized.xlsx"
    bundle = tmp_path / "medium.xlcloak"

    detector = PiiDetector()
    sanitizer = Sanitizer(detector)
    result = sanitizer.run(medium_fixture, output_path=output, bundle_path=bundle, hide_all=True)

    assert result.sanitized_path.exists(), "Sanitized file must be written"
    assert result.bundle_path.exists(), "Bundle must be written"
    assert result.cells_sanitized > 0, "hide-all must have replaced at least one cell"
    assert result.token_count > 0, "At least one unique token must be registered"
