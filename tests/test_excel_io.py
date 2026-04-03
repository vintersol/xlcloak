"""Tests for Excel I/O pipeline — WorkbookReader and WorkbookWriter."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment

from xlcloak.excel_io import WorkbookReader, WorkbookWriter


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_workbook(tmp_path: Path, name: str = "test.xlsx") -> Path:
    """Create an empty workbook and return its path."""
    wb = Workbook()
    path = tmp_path / name
    wb.save(str(path))
    return path


def _make_text_workbook(tmp_path: Path, name: str = "test.xlsx") -> Path:
    """Create a workbook with text cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Alice"
    ws["B1"] = "bob@example.com"
    ws["A2"] = "Charlie"
    path = tmp_path / name
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# Round-trip tests
# ---------------------------------------------------------------------------


def test_roundtrip_text_values(tmp_path: Path) -> None:
    """Text cells survive a read/write round-trip without modification."""
    src = _make_text_workbook(tmp_path, "source.xlsx")
    out = tmp_path / "output.xlsx"

    reader = WorkbookReader(src)
    wb = reader.open()
    cells = list(reader.iter_text_cells(wb))

    writer = WorkbookWriter(src, out)
    patches = [(c.sheet_name, c.row, c.col, c.value) for c in cells]
    writer.patch_and_save(patches)

    # Read back and verify values
    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2["Sheet1"]
    assert ws2["A1"].value == "Alice"
    assert ws2["B1"].value == "bob@example.com"
    assert ws2["A2"].value == "Charlie"


def test_roundtrip_preserves_numbers(tmp_path: Path) -> None:
    """Numeric cells are not affected when patching only text cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Hello"
    ws["B1"] = 42
    ws["C1"] = 3.14
    src = tmp_path / "source.xlsx"
    wb.save(str(src))
    out = tmp_path / "output.xlsx"

    reader = WorkbookReader(src)
    wb_in = reader.open()
    cells = list(reader.iter_text_cells(wb_in))

    writer = WorkbookWriter(src, out)
    patches = [(c.sheet_name, c.row, c.col, "PATCHED") for c in cells]
    writer.patch_and_save(patches)

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2["Sheet1"]
    assert ws2["A1"].value == "PATCHED"
    assert ws2["B1"].value == 42
    assert ws2["C1"].value == pytest.approx(3.14)


def test_roundtrip_preserves_formatting(tmp_path: Path) -> None:
    """Bold formatting is preserved after patching cell values."""
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "BoldText"
    ws["A1"].font = Font(bold=True)
    src = tmp_path / "source.xlsx"
    wb.save(str(src))
    out = tmp_path / "output.xlsx"

    reader = WorkbookReader(src)
    wb_in = reader.open()
    cells = list(reader.iter_text_cells(wb_in))

    writer = WorkbookWriter(src, out)
    patches = [(c.sheet_name, c.row, c.col, "NewText") for c in cells]
    writer.patch_and_save(patches)

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2["Sheet1"]
    assert ws2["A1"].value == "NewText"
    assert ws2["A1"].font.bold is True


def test_patch_replaces_value(tmp_path: Path) -> None:
    """patch_and_save replaces the target cell and leaves others unchanged."""
    src = _make_text_workbook(tmp_path, "source.xlsx")
    out = tmp_path / "output.xlsx"

    writer = WorkbookWriter(src, out)
    writer.patch_and_save([("Sheet1", 1, 1, "REPLACED")])

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2["Sheet1"]
    assert ws2.cell(row=1, column=1).value == "REPLACED"
    assert ws2.cell(row=1, column=2).value == "bob@example.com"
    assert ws2.cell(row=2, column=1).value == "Charlie"


def test_iter_text_cells_skips_numbers(tmp_path: Path) -> None:
    """iter_text_cells does not yield numeric cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "text"
    ws["B1"] = 100
    ws["C1"] = 3.14
    src = tmp_path / "source.xlsx"
    wb.save(str(src))

    reader = WorkbookReader(src)
    wb_in = reader.open()
    cells = list(reader.iter_text_cells(wb_in))

    sheet_cells = [(c.row, c.col) for c in cells if c.sheet_name == "Sheet1"]
    assert (1, 1) in sheet_cells   # "text" should be included
    assert (1, 2) not in sheet_cells  # 100 excluded
    assert (1, 3) not in sheet_cells  # 3.14 excluded


def test_iter_text_cells_skips_none(tmp_path: Path) -> None:
    """iter_text_cells does not yield empty (None value) cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "present"
    # B1 and C1 are empty — should not be yielded
    src = tmp_path / "source.xlsx"
    wb.save(str(src))

    reader = WorkbookReader(src)
    wb_in = reader.open()
    cells = list(reader.iter_text_cells(wb_in))

    # Only the one text cell should be present
    assert len(cells) == 1
    assert cells[0].value == "present"


# ---------------------------------------------------------------------------
# Surface detection tests
# ---------------------------------------------------------------------------


def test_detect_formula(tmp_path: Path) -> None:
    """Formula cells are reported as WARNING-level SurfaceWarnings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=SUM(A1:A2)"
    src = tmp_path / "formula.xlsx"
    wb.save(str(src))

    reader = WorkbookReader(src)
    wb_in = reader.open()
    warnings = reader.scan_surfaces(wb_in)

    formula_warnings = [w for w in warnings if w.surface_type == "formula"]
    assert len(formula_warnings) >= 1
    assert formula_warnings[0].level == "WARNING"


def test_detect_comment(tmp_path: Path) -> None:
    """Cells with comments are reported as WARNING-level SurfaceWarnings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "data"
    ws["A1"].comment = Comment("this is a note", "author")
    src = tmp_path / "comment.xlsx"
    wb.save(str(src))

    reader = WorkbookReader(src)
    wb_in = reader.open()
    warnings = reader.scan_surfaces(wb_in)

    comment_warnings = [w for w in warnings if w.surface_type == "comment"]
    assert len(comment_warnings) >= 1
    assert comment_warnings[0].level == "WARNING"
    assert "this is a note" in comment_warnings[0].detail


def test_detect_merged_cells(tmp_path: Path) -> None:
    """Merged cell ranges are reported as INFO-level SurfaceWarnings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.merge_cells("A1:B2")
    src = tmp_path / "merged.xlsx"
    wb.save(str(src))

    reader = WorkbookReader(src)
    wb_in = reader.open()
    warnings = reader.scan_surfaces(wb_in)

    merged_warnings = [w for w in warnings if w.surface_type == "merged_cells"]
    assert len(merged_warnings) >= 1
    assert merged_warnings[0].level == "INFO"


def test_detect_no_surfaces(tmp_path: Path) -> None:
    """A clean workbook with only text cells returns an empty warnings list."""
    src = _make_text_workbook(tmp_path, "clean.xlsx")

    reader = WorkbookReader(src)
    wb_in = reader.open()
    warnings = reader.scan_surfaces(wb_in)

    assert warnings == []


def test_multiple_surfaces(tmp_path: Path) -> None:
    """A workbook with both a formula and a comment produces 2+ warnings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=SUM(1,2)"
    ws["B1"] = "data"
    ws["B1"].comment = Comment("note", "author")
    src = tmp_path / "multi.xlsx"
    wb.save(str(src))

    reader = WorkbookReader(src)
    wb_in = reader.open()
    warnings = reader.scan_surfaces(wb_in)

    assert len(warnings) >= 2
    surface_types = {w.surface_type for w in warnings}
    assert "formula" in surface_types
    assert "comment" in surface_types


# NOTE: Chart and image detection cannot be easily tested with openpyxl's
# programmatic API. These surfaces are validated via hard fixture files in Plan 03.
