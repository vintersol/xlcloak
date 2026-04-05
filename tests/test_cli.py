"""CLI integration tests for the xlcloak sanitize command."""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
import pytest
import spacy.util
from click.testing import CliRunner

from xlcloak.bundle import BundleWriter, DEFAULT_PASSWORD
from xlcloak.cli import main
from xlcloak.sanitizer import derive_output_paths

# Mark for tests that require the spaCy model
requires_spacy = pytest.mark.skipif(
    not spacy.util.is_package("en_core_web_lg"),
    reason="spaCy model en_core_web_lg not installed",
)

# Apply to entire module — all tests here need spaCy for entity detection
# (except test_package_install_and_help and test_package_version which are
# defined without the module mark by using pytestmark override at the module level;
# those tests are not affected since --help/--version don't load the spaCy model,
# but pytest module-level mark applies to all functions — acceptable: they skip too)
pytestmark = requires_spacy

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

FIXTURES_DIR = Path(__file__).parent / "fixtures"
SIMPLE_FIXTURE = FIXTURES_DIR / "simple.xlsx"


# ---------------------------------------------------------------------------
# CLI tests
# ---------------------------------------------------------------------------


def test_cli_sanitize_produces_outputs(tmp_path: Path) -> None:
    """sanitize command exits 0 and produces three output files."""
    fixture = tmp_path / "simple.xlsx"
    shutil.copy2(SIMPLE_FIXTURE, fixture)

    runner = CliRunner()
    result = runner.invoke(main, ["sanitize", str(fixture)])

    assert result.exit_code == 0, f"Expected exit 0, got {result.exit_code}: {result.output}"

    sanitized, bundle, manifest = derive_output_paths(fixture)
    assert sanitized.exists(), f"Sanitized file not found: {sanitized}"
    assert bundle.exists(), f"Bundle file not found: {bundle}"
    assert manifest.exists(), f"Manifest file not found: {manifest}"


def test_cli_sanitize_overwrite_protection(tmp_path: Path) -> None:
    """Second invocation without --force exits non-zero with --force hint."""
    fixture = tmp_path / "simple.xlsx"
    shutil.copy2(SIMPLE_FIXTURE, fixture)

    runner = CliRunner()
    runner.invoke(main, ["sanitize", str(fixture)])  # first run

    result = runner.invoke(main, ["sanitize", str(fixture)])  # second run
    assert result.exit_code != 0, "Expected non-zero exit on overwrite without --force"
    assert "--force" in result.output, f"Expected --force hint in output: {result.output!r}"


def test_cli_sanitize_output_flag(tmp_path: Path) -> None:
    """--output flag redirects sanitized file to specified path."""
    fixture = tmp_path / "simple.xlsx"
    shutil.copy2(SIMPLE_FIXTURE, fixture)

    custom_output = tmp_path / "out" / "custom.xlsx"
    custom_output.parent.mkdir(parents=True, exist_ok=True)

    runner = CliRunner()
    result = runner.invoke(main, ["sanitize", str(fixture), "--output", str(custom_output)])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    sanitized, bundle, manifest = derive_output_paths(fixture, custom_output)
    assert sanitized.exists(), f"Sanitized file not found: {sanitized}"
    assert bundle.exists(), f"Bundle file not found: {bundle}"
    assert manifest.exists(), f"Manifest file not found: {manifest}"


def test_cli_sanitize_default_password_warning(tmp_path: Path) -> None:
    """Invoking sanitize without --password shows default password warning."""
    fixture = tmp_path / "simple.xlsx"
    shutil.copy2(SIMPLE_FIXTURE, fixture)

    runner = CliRunner()
    result = runner.invoke(main, ["sanitize", str(fixture)])

    assert "default password" in result.stderr, (
        f"Expected 'default password' in stderr: {result.stderr!r}"
    )


# ---------------------------------------------------------------------------
# Inspect command tests
# ---------------------------------------------------------------------------


def test_inspect_shows_summary() -> None:
    """inspect on simple fixture exits 0 and output contains entity type names."""
    runner = CliRunner()
    result = runner.invoke(main, ["inspect", str(SIMPLE_FIXTURE)])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    # At least one entity type should appear in the output
    found = any(
        etype in result.output
        for etype in ("EMAIL", "PERSON", "PHONE", "URL", "ORG")
    )
    assert found, f"Expected at least one entity type in output: {result.output!r}"


def test_inspect_no_output_files(tmp_path: Path) -> None:
    """inspect does not write any files to the directory."""
    fixture = tmp_path / "simple.xlsx"
    shutil.copy2(SIMPLE_FIXTURE, fixture)

    files_before = set(tmp_path.iterdir())
    runner = CliRunner()
    runner.invoke(main, ["inspect", str(fixture)])
    files_after = set(tmp_path.iterdir())

    new_files = files_after - files_before
    assert not new_files, f"Expected no new files, but found: {new_files}"


def test_inspect_shows_table() -> None:
    """inspect output contains table column headers Sheet, Cell, Type."""
    runner = CliRunner()
    result = runner.invoke(main, ["inspect", str(SIMPLE_FIXTURE)])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    assert "Sheet" in result.output, f"Expected 'Sheet' in output: {result.output!r}"
    assert "Cell" in result.output, f"Expected 'Cell' in output: {result.output!r}"
    assert "Type" in result.output, f"Expected 'Type' in output: {result.output!r}"


def test_inspect_shows_warnings() -> None:
    """inspect on hard fixture (has formulas) shows warnings section."""
    hard_fixture = FIXTURES_DIR / "hard.xlsx"
    runner = CliRunner()
    result = runner.invoke(main, ["inspect", str(hard_fixture)])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    # Hard fixture has formulas; should show warnings
    assert "Warning" in result.output or "formula" in result.output, (
        f"Expected 'Warning' or 'formula' in output: {result.output!r}"
    )


def test_inspect_verbose() -> None:
    """inspect --verbose output contains Score column header."""
    runner = CliRunner()
    result = runner.invoke(main, ["inspect", str(SIMPLE_FIXTURE), "--verbose"])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    assert "Score" in result.output, f"Expected 'Score' in output: {result.output!r}"


def test_inspect_no_files_written_message() -> None:
    """inspect output contains 'No files written' message."""
    runner = CliRunner()
    result = runner.invoke(main, ["inspect", str(SIMPLE_FIXTURE)])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    assert "No files written" in result.output, (
        f"Expected 'No files written' in output: {result.output!r}"
    )


def test_package_install_and_help() -> None:
    """xlcloak --help exits 0 and shows sanitize and inspect commands."""
    runner = CliRunner()
    result = runner.invoke(main, ["--help"])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    assert "sanitize" in result.output, f"Expected 'sanitize' in output: {result.output!r}"
    assert "inspect" in result.output, f"Expected 'inspect' in output: {result.output!r}"


def test_package_version() -> None:
    """xlcloak --version exits 0 and shows version string."""
    runner = CliRunner()
    result = runner.invoke(main, ["--version"])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    assert "0.1.0" in result.output, f"Expected '0.1.0' in output: {result.output!r}"


# ---------------------------------------------------------------------------
# Restore command helpers
# ---------------------------------------------------------------------------


def _make_sanitized_xlsx_and_bundle(tmp_path: Path) -> tuple[Path, Path]:
    """Create a minimal sanitized xlsx (with tokens) and matching bundle."""
    # Build sanitized xlsx with tokens
    sanitized = tmp_path / "data_sanitized.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1).value = "PERSON_001"
    ws.cell(row=2, column=1).value = "EMAIL_002@example.com"
    wb.save(str(sanitized))

    # Write matching bundle
    forward_map = {
        "John Smith": "PERSON_001",
        "john@example.com": "EMAIL_002@example.com",
    }
    reverse_map = {v: k for k, v in forward_map.items()}
    bundle = tmp_path / "data.xlcloak"
    BundleWriter(DEFAULT_PASSWORD).write(
        bundle,
        forward_map=forward_map,
        reverse_map=reverse_map,
        original_filename="data.xlsx",
        sheets_processed=["Sheet1"],
        token_count=len(forward_map),
    )
    return sanitized, bundle


# ---------------------------------------------------------------------------
# Restore command tests
# ---------------------------------------------------------------------------


def test_restore_help() -> None:
    """xlcloak restore --help exits 0 and shows required options."""
    runner = CliRunner()
    result = runner.invoke(main, ["restore", "--help"])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    assert "--bundle" in result.output, f"Expected '--bundle' in output: {result.output!r}"
    assert "--password" in result.output, f"Expected '--password' in output: {result.output!r}"
    assert "--force" in result.output, f"Expected '--force' in output: {result.output!r}"
    assert "--verbose" in result.output, f"Expected '--verbose' in output: {result.output!r}"
    assert "--output" in result.output, f"Expected '--output' in output: {result.output!r}"


def test_restore_produces_restored_file(tmp_path: Path) -> None:
    """restore command exits 0 and produces _restored.xlsx file."""
    sanitized, bundle = _make_sanitized_xlsx_and_bundle(tmp_path)

    runner = CliRunner()
    result = runner.invoke(
        main,
        ["restore", str(sanitized), "--bundle", str(bundle), "--force"],
    )

    assert result.exit_code == 0, f"Expected exit 0: {result.output}\n{result.exception}"
    restored = tmp_path / "data_sanitized_restored.xlsx"
    assert restored.exists(), f"Restored file not found: {restored}"
    assert "Cells restored:" in result.output


def test_restore_wrong_password_exits_error(tmp_path: Path) -> None:
    """restore with wrong password exits non-zero and shows 'Invalid password' error."""
    sanitized, bundle = _make_sanitized_xlsx_and_bundle(tmp_path)

    runner = CliRunner()
    result = runner.invoke(
        main,
        ["restore", str(sanitized), "--bundle", str(bundle), "--password", "wrong_pw"],
    )

    assert result.exit_code != 0, f"Expected non-zero exit, got {result.exit_code}"
    assert "Invalid password" in result.output or "Invalid password" in (result.stderr or ""), (
        f"Expected 'Invalid password' in output: {result.output!r}"
    )


def test_restore_overwrite_protection(tmp_path: Path) -> None:
    """Second restore without --force exits non-zero with --force hint."""
    sanitized, bundle = _make_sanitized_xlsx_and_bundle(tmp_path)

    runner = CliRunner()
    runner.invoke(main, ["restore", str(sanitized), "--bundle", str(bundle), "--force"])

    result = runner.invoke(main, ["restore", str(sanitized), "--bundle", str(bundle)])
    assert result.exit_code != 0, "Expected non-zero exit on overwrite without --force"
    assert "--force" in result.output, f"Expected '--force' hint in output: {result.output!r}"


# ---------------------------------------------------------------------------
# diff command tests
# (These tests do not require the spaCy model — override module-level mark)
# ---------------------------------------------------------------------------

# Marker that always runs (counterpart to the module-level requires_spacy mark)
no_spacy_needed = pytest.mark.skipif(False, reason="no spaCy needed")


@no_spacy_needed
def test_diff_help() -> None:
    """xlcloak diff --help exits 0 and shows --bundle and --verbose options."""
    runner = CliRunner()
    result = runner.invoke(main, ["diff", "--help"])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    assert "--bundle" in result.output, f"Expected '--bundle' in output: {result.output!r}"
    assert "--verbose" in result.output, f"Expected '--verbose' in output: {result.output!r}"
    assert "--password" in result.output, f"Expected '--password' in output: {result.output!r}"


def test_diff_no_changes(tmp_path: Path) -> None:
    """diff with no AI changes shows 'No tokens changed by AI.' and 'No files written.'."""
    sanitized, bundle = _make_sanitized_xlsx_and_bundle(tmp_path)

    runner = CliRunner()
    result = runner.invoke(
        main, ["diff", str(sanitized), "--bundle", str(bundle)]
    )

    assert result.exit_code == 0, f"Expected exit 0: {result.output}\n{result.exception}"
    assert "No tokens changed by AI." in result.output, (
        f"Expected 'No tokens changed by AI.' in output: {result.output!r}"
    )
    assert "No files written." in result.output, (
        f"Expected 'No files written.' in output: {result.output!r}"
    )


def test_diff_with_ai_changes(tmp_path: Path) -> None:
    """diff with AI-modified cells shows changed token count and table."""
    # Build a sanitized xlsx where tokens have been replaced by AI (tokens are absent)
    ai_modified = tmp_path / "data_ai.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1).value = "John Smith"  # AI replaced token with original-like text
    ws.cell(row=2, column=1).value = "some other text"
    wb.save(str(ai_modified))

    # Write bundle with tokens that are no longer in the file
    forward_map = {
        "John Smith": "PERSON_001",
        "john@example.com": "EMAIL_002@example.com",
    }
    reverse_map = {v: k for k, v in forward_map.items()}
    bundle = tmp_path / "data.xlcloak"
    BundleWriter(DEFAULT_PASSWORD).write(
        bundle,
        forward_map=forward_map,
        reverse_map=reverse_map,
        original_filename="data.xlsx",
        sheets_processed=["Sheet1"],
        token_count=len(forward_map),
    )

    runner = CliRunner()
    result = runner.invoke(
        main, ["diff", str(ai_modified), "--bundle", str(bundle)]
    )

    assert result.exit_code == 0, f"Expected exit 0: {result.output}\n{result.exception}"
    assert "token(s) changed by AI" in result.output or "changed by AI" in result.output, (
        f"Expected changed-token message in output: {result.output!r}"
    )
    assert "No files written." in result.output, (
        f"Expected 'No files written.' in output: {result.output!r}"
    )


def test_diff_verbose_shows_unchanged(tmp_path: Path) -> None:
    """diff --verbose shows unchanged tokens in addition to changed ones."""
    # Build a sanitized xlsx where ONE token is still there, one is missing
    partial_sanitized = tmp_path / "partial_sanitized.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1).value = "PERSON_001"  # still in file (unchanged)
    # EMAIL_002@example.com is missing (AI changed it)
    wb.save(str(partial_sanitized))

    forward_map = {
        "John Smith": "PERSON_001",
        "john@example.com": "EMAIL_002@example.com",
    }
    reverse_map = {v: k for k, v in forward_map.items()}
    bundle = tmp_path / "data.xlcloak"
    BundleWriter(DEFAULT_PASSWORD).write(
        bundle,
        forward_map=forward_map,
        reverse_map=reverse_map,
        original_filename="data.xlsx",
        sheets_processed=["Sheet1"],
        token_count=len(forward_map),
    )

    runner = CliRunner()
    result = runner.invoke(
        main, ["diff", str(partial_sanitized), "--bundle", str(bundle), "--verbose"]
    )

    assert result.exit_code == 0, f"Expected exit 0: {result.output}\n{result.exception}"
    assert "Unchanged tokens" in result.output, (
        f"Expected 'Unchanged tokens' in output: {result.output!r}"
    )
    assert "No files written." in result.output, (
        f"Expected 'No files written.' in output: {result.output!r}"
    )


def test_diff_detects_partial_loss_of_duplicate_token_occurrences(tmp_path: Path) -> None:
    """diff should report missing occurrences when only some duplicate token cells remain."""
    partial_sanitized = tmp_path / "partial_sanitized.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1).value = "PERSON_001"
    ws.cell(row=2, column=1).value = "AI changed this"
    wb.save(str(partial_sanitized))

    bundle = tmp_path / "data.xlcloak"
    BundleWriter(DEFAULT_PASSWORD).write(
        bundle,
        forward_map={"John Smith": "PERSON_001"},
        reverse_map={"PERSON_001": "John Smith"},
        original_filename="data.xlsx",
        sheets_processed=["Sheet1"],
        token_count=1,
        token_occurrences={"PERSON_001": 2},
    )

    runner = CliRunner()
    result = runner.invoke(
        main, ["diff", str(partial_sanitized), "--bundle", str(bundle)]
    )

    assert result.exit_code == 0, f"Expected exit 0: {result.output}\n{result.exception}"
    assert "1 token occurrence(s) changed by AI." in result.output
    assert "PERSON_001" in result.output
    assert "Missing Occurrences" in result.output


def test_diff_wrong_password(tmp_path: Path) -> None:
    """diff with wrong password exits non-zero and shows error."""
    sanitized, bundle = _make_sanitized_xlsx_and_bundle(tmp_path)

    runner = CliRunner()
    result = runner.invoke(
        main, ["diff", str(sanitized), "--bundle", str(bundle), "--password", "wrong_pw"]
    )

    assert result.exit_code != 0, f"Expected non-zero exit: {result.output}"
    assert "Error" in result.output or "Error" in (result.stderr or ""), (
        f"Expected 'Error' in output or stderr: {result.output!r}"
    )


def test_diff_no_files_written(tmp_path: Path) -> None:
    """diff command writes no files to disk."""
    sanitized, bundle = _make_sanitized_xlsx_and_bundle(tmp_path)
    files_before = set(tmp_path.iterdir())

    runner = CliRunner()
    runner.invoke(main, ["diff", str(sanitized), "--bundle", str(bundle)])

    files_after = set(tmp_path.iterdir())
    new_files = files_after - files_before
    assert not new_files, f"Expected no new files, but found: {new_files}"


# ---------------------------------------------------------------------------
# CLI alias tests
# (reconcile -> restore, deidentify -> sanitize, identify -> restore)
# These tests do not require the spaCy model.
# ---------------------------------------------------------------------------


@no_spacy_needed
def test_alias_reconcile_help() -> None:
    """xlcloak reconcile --help exits 0 and shows restore options."""
    runner = CliRunner()
    result = runner.invoke(main, ["reconcile", "--help"])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    assert "--bundle" in result.output, f"Expected '--bundle' in output: {result.output!r}"
    assert "--password" in result.output, f"Expected '--password' in output: {result.output!r}"
    assert "--force" in result.output, f"Expected '--force' in output: {result.output!r}"


@no_spacy_needed
def test_alias_deidentify_help() -> None:
    """xlcloak deidentify --help exits 0 and shows sanitize options."""
    runner = CliRunner()
    result = runner.invoke(main, ["deidentify", "--help"])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    assert "--password" in result.output, f"Expected '--password' in output: {result.output!r}"


@no_spacy_needed
def test_alias_identify_help() -> None:
    """xlcloak identify --help exits 0 and shows restore options."""
    runner = CliRunner()
    result = runner.invoke(main, ["identify", "--help"])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    assert "--bundle" in result.output, f"Expected '--bundle' in output: {result.output!r}"
    assert "--password" in result.output, f"Expected '--password' in output: {result.output!r}"


@no_spacy_needed
def test_alias_reconcile_runs_restore(tmp_path: Path) -> None:
    """reconcile with valid args produces same result as restore."""
    sanitized, bundle = _make_sanitized_xlsx_and_bundle(tmp_path)

    runner = CliRunner()
    result = runner.invoke(
        main,
        ["reconcile", str(sanitized), "--bundle", str(bundle), "--force"],
    )

    assert result.exit_code == 0, f"Expected exit 0: {result.output}\n{result.exception}"
    assert "Restored:" in result.output or "Cells restored:" in result.output, (
        f"Expected restore output: {result.output!r}"
    )


@no_spacy_needed
def test_main_help_lists_all_commands() -> None:
    """xlcloak --help lists all commands including aliases."""
    runner = CliRunner()
    result = runner.invoke(main, ["--help"])

    assert result.exit_code == 0, f"Expected exit 0: {result.output}"
    assert "reconcile" in result.output, f"Expected 'reconcile' in output: {result.output!r}"
    assert "deidentify" in result.output, f"Expected 'deidentify' in output: {result.output!r}"
    assert "identify" in result.output, f"Expected 'identify' in output: {result.output!r}"


# ---------------------------------------------------------------------------
# --hide-all flag tests
# ---------------------------------------------------------------------------


@no_spacy_needed
def test_cli_hide_all_flag_in_help() -> None:
    """--hide-all flag appears in xlcloak sanitize --help."""
    from xlcloak.cli import sanitize
    runner = CliRunner()
    result = runner.invoke(sanitize, ["--help"])
    assert "--hide-all" in result.output, f"Expected '--hide-all' in help: {result.output!r}"


def test_cli_hide_all_dry_run(simple_fixture, tmp_path) -> None:
    """--hide-all --dry-run prints cell count and writes no files."""
    runner = CliRunner()
    result = runner.invoke(main, ["sanitize", str(simple_fixture), "--hide-all", "--dry-run"])
    assert result.exit_code == 0, result.output
    assert "hide-all" in result.output.lower(), (
        f"Expected 'hide-all' in output: {result.output!r}"
    )
    assert "Would replace" in result.output, (
        f"Expected 'Would replace' in output: {result.output!r}"
    )
