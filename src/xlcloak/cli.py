"""Click CLI entry point for xlcloak."""

from __future__ import annotations

import re
import sys
from collections import Counter
from pathlib import Path

import click
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.table import Table

import xlcloak
from xlcloak.bundle import DEFAULT_PASSWORD


@click.group(context_settings={"auto_envvar_prefix": "XLCLOAK"})
@click.version_option(version=xlcloak.__version__, prog_name="xlcloak")
def main() -> None:
    """xlcloak -- reversible Excel text sanitization for AI workflows."""


@main.command()
@click.argument("file", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--password",
    default=DEFAULT_PASSWORD,
    show_default=True,
    help="Encryption password for the restore bundle",
)
@click.option(
    "--output",
    "output_path",
    type=click.Path(path_type=Path),
    default=None,
    help="Output path for sanitized file (bundle and manifest derive from it)",
)
@click.option(
    "--bundle",
    "bundle_path",
    type=click.Path(path_type=Path),
    default=None,
    help="Explicit output path for the encrypted bundle (overrides default naming)",
)
@click.option(
    "--dry-run",
    is_flag=True,
    default=False,
    help="Preview what would be sanitized without writing any files",
)
@click.option(
    "--text-mode",
    is_flag=True,
    default=False,
    help="Extract text cells to a plain text file without token replacement",
)
@click.option(
    "--force",
    is_flag=True,
    default=False,
    help="Overwrite existing output files",
)
@click.option(
    "--hide-all",
    is_flag=True,
    default=False,
    help="Replace every text cell with a stable token regardless of content",
)
@click.option(
    "--allow-unsupported-surfaces",
    is_flag=True,
    default=False,
    help="Deprecated compatibility flag; currently has no effect",
)
@click.option(
    "--verbose",
    is_flag=True,
    default=False,
    help="Show detailed output",
)
def sanitize(
    file: Path,
    password: str,
    output_path: Path | None,
    bundle_path: Path | None,
    dry_run: bool,
    text_mode: bool,
    force: bool,
    hide_all: bool,
    allow_unsupported_surfaces: bool,
    verbose: bool,
) -> None:
    """Sanitize FILE, producing a sanitized xlsx, encrypted bundle, and manifest."""
    from xlcloak.detector import PiiDetector
    from xlcloak.excel_io import WorkbookReader
    from xlcloak.sanitizer import Sanitizer, derive_output_paths
    from xlcloak.token_engine import TokenRegistry

    if dry_run:
        if hide_all:
            try:
                reader = WorkbookReader(file)
                wb = reader.open()
                n = sum(1 for _ in reader.iter_text_cells(wb))
            except Exception as exc:
                click.echo(f"Error: {exc}", err=True)
                sys.exit(1)
            click.echo(f"Dry run (hide-all): Would replace {n} text cells.")
            click.echo("No files written.")
            return
        try:
            detector = PiiDetector()
            registry = TokenRegistry()
            reader = WorkbookReader(file)
            wb = reader.open()
            # Pre-pass: extract column headers (mirrors Sanitizer.run)
            text_cells = list(reader.iter_text_cells(wb))
            sheet_headers: dict[str, dict[int, str]] = {}
            for cell_ref in text_cells:
                if cell_ref.row == 1:
                    sheet_headers.setdefault(cell_ref.sheet_name, {})[cell_ref.col] = cell_ref.value or ""
            all_results = []
            for cell_ref in text_cells:
                col_header = (
                    sheet_headers.get(cell_ref.sheet_name, {}).get(cell_ref.col)
                    if cell_ref.row > 1
                    else None
                )
                scan_results, _replaced = detector.detect_cell(cell_ref, registry, column_header=col_header)
                all_results.extend(scan_results)
        except Exception as exc:
            click.echo(f"Error: {exc}", err=True)
            sys.exit(1)

        click.echo(f"Dry run: {file.name}")
        if all_results:
            entity_counts: dict[str, int] = {}
            for r in all_results:
                entity_counts[r.entity_type.value] = entity_counts.get(r.entity_type.value, 0) + 1
            click.echo(f"Would replace {len(registry)} unique entities across {len(all_results)} detections:")
            for etype, count in sorted(entity_counts.items()):
                click.echo(f"  {etype}: {count}")
        else:
            click.echo("No entities detected — no changes would be made.")
        click.echo("No files written.")
        return

    if text_mode:
        try:
            reader = WorkbookReader(file)
            wb = reader.open()
            text_cells = list(reader.iter_text_cells(wb))
        except Exception as exc:
            click.echo(f"Error: {exc}", err=True)
            sys.exit(1)

        _, _, base_manifest = derive_output_paths(file, output_path)
        text_out = base_manifest.parent / (base_manifest.stem.replace("_manifest", "_text") + ".txt")
        if text_out.exists() and not force:
            click.echo(f"Error: {text_out} already exists. Use --force to overwrite.", err=True)
            sys.exit(1)
        lines = ["sheet\tcell\tvalue"]
        for cell_ref in text_cells:
            col_letter = get_column_letter(cell_ref.col)
            lines.append(f"{cell_ref.sheet_name}\t{col_letter}{cell_ref.row}\t{cell_ref.value}")
        text_out.write_text("\n".join(lines))
        click.echo(f"Text extracted: {text_out} ({len(text_cells)} cells)")
        return

    if password == DEFAULT_PASSWORD:
        click.echo(
            "Warning: Using default password. Use --password for real encryption.",
            err=True,
        )
    if allow_unsupported_surfaces:
        click.echo(
            "Warning: --allow-unsupported-surfaces is deprecated and has no effect.",
            err=True,
        )

    try:
        detector = PiiDetector()
        sanitizer = Sanitizer(detector, password)
        result = sanitizer.run(file, output_path, force, bundle_path, hide_all=hide_all)
    except click.UsageError:
        raise
    except Exception as exc:
        click.echo(f"Error: {exc}", err=True)
        sys.exit(1)

    click.echo(f"Sanitized: {result.sanitized_path}")
    click.echo(f"Bundle:    {result.bundle_path}")
    click.echo(f"Manifest:  {result.manifest_path}")
    click.echo(f"Tokens:    {result.token_count} unique entities replaced")

    if verbose and result.entity_counts:
        click.echo("")
        click.echo("Entity breakdown:")
        for entity_type, count in sorted(result.entity_counts.items()):
            click.echo(f"  {entity_type}: {count}")


@main.command()
@click.argument("file", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--bundle",
    "bundle_path",
    required=True,
    type=click.Path(exists=True, path_type=Path),
    help="Path to the .xlcloak restore bundle",
)
@click.option(
    "--password",
    default=DEFAULT_PASSWORD,
    show_default=True,
    help="Decryption password for the bundle",
)
@click.option(
    "--output",
    "output_path",
    type=click.Path(path_type=Path),
    default=None,
    help="Output path for restored file",
)
@click.option(
    "--force",
    is_flag=True,
    default=False,
    help="Overwrite existing output files",
)
@click.option(
    "--verbose",
    is_flag=True,
    default=False,
    help="Show detailed output including skipped token list",
)
def restore(
    file: Path,
    bundle_path: Path,
    password: str,
    output_path: Path | None,
    force: bool,
    verbose: bool,
) -> None:
    """Restore FILE from a sanitized xlsx using the encrypted BUNDLE."""
    from xlcloak.restorer import Restorer

    try:
        result = Restorer(password).run(file, bundle_path, output_path, force)
    except ValueError as exc:
        click.echo(f"Error: {exc}", err=True)
        sys.exit(1)
    except click.UsageError:
        raise
    except Exception as exc:
        click.echo(f"Error: {exc}", err=True)
        sys.exit(1)

    click.echo(f"Restored: {result.restored_path}")
    click.echo(f"Manifest: {result.manifest_path}")
    click.echo(f"Cells restored: {result.restored_count}")

    if result.skipped_count > 0:
        click.echo(f"Skipped (AI-modified): {result.skipped_count}")

    if verbose and result.skipped_cells:
        click.echo("")
        click.echo("Skipped tokens:")
        for sc in result.skipped_cells:
            count = int(sc.get("count", 1))
            if count > 1:
                click.echo(f"  {sc['token']} (was: {sc['original']}) x{count}")
            else:
                click.echo(f"  {sc['token']} (was: {sc['original']})")


@main.command()
@click.argument("file", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--verbose",
    is_flag=True,
    default=False,
    help="Show confidence scores and detection method",
)
def inspect(file: Path, verbose: bool) -> None:
    """Preview what sanitize would do — no files written."""
    from xlcloak.detector import PiiDetector
    from xlcloak.excel_io import WorkbookReader
    from xlcloak.token_engine import TokenRegistry

    try:
        detector = PiiDetector()
        registry = TokenRegistry()
        reader = WorkbookReader(file)
        wb = reader.open()

        text_cells = list(reader.iter_text_cells(wb))
        sheet_headers: dict[str, dict[int, str]] = {}
        for cell_ref in text_cells:
            if cell_ref.row == 1:
                sheet_headers.setdefault(cell_ref.sheet_name, {})[cell_ref.col] = cell_ref.value or ""

        all_results = []
        for cell_ref in text_cells:
            col_header = (
                sheet_headers.get(cell_ref.sheet_name, {}).get(cell_ref.col)
                if cell_ref.row > 1
                else None
            )
            scan_results, _replaced = detector.detect_cell(cell_ref, registry, column_header=col_header)
            all_results.extend(scan_results)

        warnings = reader.scan_surfaces(wb)
    except Exception as exc:
        click.echo(f"Error: {exc}", err=True)
        sys.exit(1)

    # Summary header
    click.echo(f"xlcloak inspect: {file.name}")
    click.echo("")

    if all_results:
        entity_counts: dict[str, int] = {}
        for r in all_results:
            key = r.entity_type.value
            entity_counts[key] = entity_counts.get(key, 0) + 1

        click.echo(f"Detected {len(all_results)} entities:")
        for etype, count in sorted(entity_counts.items()):
            click.echo(f"  {etype}: {count}")
        click.echo("")

        # Per-cell table
        console = Console()
        table = Table(show_header=True, header_style="bold")
        table.add_column("Sheet")
        table.add_column("Cell")
        table.add_column("Type")
        table.add_column("Original")
        table.add_column("Would-be Token")
        if verbose:
            table.add_column("Score")
            table.add_column("Method")

        for r in all_results:
            cell_addr = get_column_letter(r.cell.col) + str(r.cell.row)
            original_display = (
                r.original[:40] + "..." if len(r.original) > 40 else r.original
            )
            row_values = [
                r.cell.sheet_name,
                cell_addr,
                r.entity_type.value,
                original_display,
                r.token,
            ]
            if verbose:
                score_str = f"{r.score:.2f}" if r.score is not None else ""
                method_str = r.detection_method or ""
                row_values.extend([score_str, method_str])
            table.add_row(*row_values)

        console.print(table)
    else:
        click.echo("No entities detected.")
        click.echo("")

    # Warnings section
    # Filter to only surface-type warnings that indicate unsupported surfaces
    # (formulas, charts, comments — not info-level items like merged cells)
    unsupported_warnings = [
        w
        for w in warnings
        if w.surface_type in ("formula", "chart", "comment")
    ]
    if unsupported_warnings:
        click.echo("")
        click.echo("Warnings (unsupported surfaces -- not sanitized):")
        for w in unsupported_warnings:
            if w.cell.row == 0 and w.cell.col == 0:
                # Sheet-level warning
                click.echo(f"  - {w.cell.sheet_name}: {w.detail}")
            else:
                cell_addr = get_column_letter(w.cell.col) + str(w.cell.row)
                click.echo(f"  - {w.cell.sheet_name}!{cell_addr}: {w.surface_type} {w.detail}")

    click.echo("")
    click.echo("No files written.")


@main.command()
@click.argument("file", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--bundle",
    "bundle_path",
    required=True,
    type=click.Path(exists=True, path_type=Path),
    help="Path to the .xlcloak restore bundle",
)
@click.option(
    "--password",
    default=DEFAULT_PASSWORD,
    show_default=True,
    help="Decryption password for the bundle",
)
@click.option(
    "--verbose",
    is_flag=True,
    default=False,
    help="Also show unchanged token cells and new cells",
)
def diff(file: Path, bundle_path: Path, password: str, verbose: bool) -> None:
    """Show what changed between a sanitized FILE and its BUNDLE — no files written."""
    from xlcloak.bundle import BundleReader
    from xlcloak.excel_io import WorkbookReader

    try:
        payload = BundleReader(password).read(bundle_path)
    except ValueError as exc:
        click.echo(f"Error: {exc}", err=True)
        sys.exit(1)
    except Exception as exc:
        click.echo(f"Error: {exc}", err=True)
        sys.exit(1)

    reverse_map: dict[str, str] = payload.get("reverse_map", {})
    raw_occurrences = payload.get("token_occurrences", {})
    if isinstance(raw_occurrences, dict) and raw_occurrences:
        expected_occurrences = {
            token: int(count)
            for token, count in raw_occurrences.items()
            if token in reverse_map and int(count) > 0
        }
        if not expected_occurrences:
            expected_occurrences = {token: 1 for token in reverse_map}
    else:
        expected_occurrences = {token: 1 for token in reverse_map}

    # Build compiled regex from reverse_map keys (same approach as restorer.py)
    if reverse_map:
        sorted_keys = sorted(reverse_map.keys(), key=len, reverse=True)
        token_pattern = re.compile("|".join(re.escape(k) for k in sorted_keys))
    else:
        token_pattern = None

    try:
        reader = WorkbookReader(file)
        wb = reader.open()
    except Exception as exc:
        click.echo(f"Error: {exc}", err=True)
        sys.exit(1)

    # Walk all text cells; classify as token (still present) or non-token
    found_tokens: dict[str, list[tuple[str, str]]] = {}  # token -> [(sheet, cell_addr), ...]
    found_occurrences: Counter[str] = Counter()
    non_token_count = 0

    for cell_ref in reader.iter_text_cells(wb):
        if token_pattern is None:
            non_token_count += 1
            continue
        matches = token_pattern.findall(cell_ref.value)
        if matches:
            cell_addr = get_column_letter(cell_ref.col) + str(cell_ref.row)
            for token in matches:
                found_occurrences[token] += 1
                found_tokens.setdefault(token, []).append(
                    (cell_ref.sheet_name, cell_addr)
                )
        else:
            non_token_count += 1

    # Token occurrences expected from sanitize but not present in the file -> AI changed them
    missing_occurrences: dict[str, int] = {}
    for token, expected_count in expected_occurrences.items():
        observed_count = found_occurrences.get(token, 0)
        if observed_count < expected_count:
            missing_occurrences[token] = expected_count - observed_count
    missing_total = sum(missing_occurrences.values())

    # Summary header
    if missing_total:
        click.echo(f"{missing_total} token occurrence(s) changed by AI.")
    else:
        click.echo("No tokens changed by AI.")

    # Table of missing (AI-modified) tokens
    if missing_occurrences:
        click.echo("")
        click.echo(
            f"{missing_total} token occurrence(s) modified by AI (missing from file):"
        )
        console = Console()
        table = Table(show_header=True, header_style="bold")
        table.add_column("Token")
        table.add_column("Original Value")
        table.add_column("Missing Occurrences")
        for token in sorted(missing_occurrences):
            table.add_row(token, reverse_map[token], str(missing_occurrences[token]))
        console.print(table)

    # Verbose: also show unchanged tokens and non-token cell count
    if verbose:
        click.echo("")
        if found_tokens:
            click.echo(f"Unchanged tokens ({len(found_tokens)}):")
            console = Console()
            vtable = Table(show_header=True, header_style="bold")
            vtable.add_column("Sheet")
            vtable.add_column("Cell")
            vtable.add_column("Token")
            vtable.add_column("Original")
            for token, locations in sorted(found_tokens.items()):
                for sheet_name, cell_addr in locations:
                    vtable.add_row(sheet_name, cell_addr, token, reverse_map[token])
            console.print(vtable)
        else:
            click.echo("Unchanged tokens (0): none")

        click.echo(f"Non-token cells: {non_token_count}")

    click.echo("")
    click.echo("No files written.")


# Aliases (per Phase 3 decisions D-01, D-02)
main.add_command(restore, name="reconcile")    # reconcile -> restore (D-01, CLI-05)
main.add_command(sanitize, name="deidentify")  # deidentify -> sanitize (D-02, CLI-07)
main.add_command(restore, name="identify")     # identify -> restore (D-02, CLI-07)
