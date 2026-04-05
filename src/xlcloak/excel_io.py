"""Excel I/O pipeline for xlcloak — read, surface scan, and copy-then-patch write."""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Iterator

import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from xlcloak.models import CellRef, SurfaceWarning

XLCLOAK_META_SHEET = "__xlcloak_meta__"
XLCLOAK_META_KEY_CELL = "A1"
XLCLOAK_META_VALUE_CELL = "B1"
XLCLOAK_BUNDLE_ID_KEY = "bundle_id"


class WorkbookReader:
    """Read a workbook, iterate text cells, and scan for unsupported surfaces."""

    def __init__(self, path: Path) -> None:
        self.path = path

    def open(self) -> Workbook:
        """Open the workbook preserving formula strings (data_only=False)."""
        return load_workbook(str(self.path), data_only=False)

    def iter_text_cells(self, wb: Workbook) -> Iterator[CellRef]:
        """Yield CellRef for every string-valued cell across all sheets."""
        for ws in wb.worksheets:
            if ws.title == XLCLOAK_META_SHEET:
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "s" and cell.value is not None:
                        yield CellRef(
                            sheet_name=ws.title,
                            row=cell.row,
                            col=cell.column,
                            value=str(cell.value),
                        )

    def scan_surfaces(self, wb: Workbook) -> list[SurfaceWarning]:
        """Scan the workbook for unsupported surfaces and return warnings."""
        warnings: list[SurfaceWarning] = []

        for ws in wb.worksheets:
            if ws.title == XLCLOAK_META_SHEET:
                continue
            # Per-cell scans
            for row in ws.iter_rows():
                for cell in row:
                    # Formulas
                    if cell.data_type == "f":
                        warnings.append(
                            SurfaceWarning(
                                cell=CellRef(
                                    sheet_name=ws.title,
                                    row=cell.row,
                                    col=cell.column,
                                ),
                                surface_type="formula",
                                detail=str(cell.value),
                                level="WARNING",
                            )
                        )
                    # Comments
                    if cell.comment is not None:
                        warnings.append(
                            SurfaceWarning(
                                cell=CellRef(
                                    sheet_name=ws.title,
                                    row=cell.row,
                                    col=cell.column,
                                ),
                                surface_type="comment",
                                detail=cell.comment.text or "",
                                level="WARNING",
                            )
                        )

            # Sheet-level scans

            # Charts
            if len(ws._charts) > 0:  # type: ignore[attr-defined]
                warnings.append(
                    SurfaceWarning(
                        cell=CellRef(sheet_name=ws.title, row=0, col=0),
                        surface_type="chart",
                        detail=f"{len(ws._charts)} chart(s)",  # type: ignore[attr-defined]
                        level="WARNING",
                    )
                )

            # Merged cells
            for merged_range in ws.merged_cells.ranges:
                warnings.append(
                    SurfaceWarning(
                        cell=CellRef(sheet_name=ws.title, row=0, col=0),
                        surface_type="merged_cells",
                        detail=str(merged_range),
                        level="INFO",
                    )
                )

            # Images
            if len(ws._images) > 0:  # type: ignore[attr-defined]
                warnings.append(
                    SurfaceWarning(
                        cell=CellRef(sheet_name=ws.title, row=0, col=0),
                        surface_type="image",
                        detail=f"{len(ws._images)} image(s)",  # type: ignore[attr-defined]
                        level="INFO",
                    )
                )

            # Data validation
            if (
                ws.data_validations
                and len(ws.data_validations.dataValidation) > 0
            ):
                warnings.append(
                    SurfaceWarning(
                        cell=CellRef(sheet_name=ws.title, row=0, col=0),
                        surface_type="data_validation",
                        detail=f"{len(ws.data_validations.dataValidation)} rule(s)",
                        level="INFO",
                    )
                )

        # Workbook-level: named ranges
        try:
            named_names = list(wb.defined_names.definedName)
            if len(named_names) > 0:
                warnings.append(
                    SurfaceWarning(
                        cell=CellRef(sheet_name="__workbook__", row=0, col=0),
                        surface_type="named_range",
                        detail=f"{len(named_names)} named range(s)",
                        level="INFO",
                    )
                )
        except AttributeError:
            # Older openpyxl uses wb.defined_names as a dict-like object
            pass

        return warnings


class WorkbookWriter:
    """Write a workbook using copy-then-patch strategy to avoid data loss."""

    def __init__(self, source_path: Path, output_path: Path) -> None:
        self.source_path = source_path
        self.output_path = output_path

    def prepare(self) -> None:
        """Copy source to output path, preserving all non-text content."""
        shutil.copy2(self.source_path, self.output_path)

    def patch_cells(self, patches: list[tuple[str, int, int, str]]) -> None:
        """Apply cell patches to the already-copied output workbook.

        Each patch is (sheet_name, row, col, new_value).
        """
        wb = openpyxl.load_workbook(str(self.output_path))
        for sheet_name, row, col, new_value in patches:
            ws = wb[sheet_name]
            ws.cell(row=row, column=col).value = new_value
        wb.save(str(self.output_path))

    def patch_and_save(self, patches: list[tuple[str, int, int, str]]) -> Path:
        """Prepare (copy) then apply patches and return output path."""
        self.prepare()
        self.patch_cells(patches)
        return self.output_path


def write_bundle_id_marker(path: Path, bundle_id: str) -> None:
    """Persist bundle_id in a very-hidden metadata sheet inside workbook at *path*."""
    wb = openpyxl.load_workbook(str(path))
    if XLCLOAK_META_SHEET in wb.sheetnames:
        ws = wb[XLCLOAK_META_SHEET]
    else:
        ws = wb.create_sheet(XLCLOAK_META_SHEET)
    ws.sheet_state = "veryHidden"
    ws[XLCLOAK_META_KEY_CELL] = XLCLOAK_BUNDLE_ID_KEY
    ws[XLCLOAK_META_VALUE_CELL] = bundle_id
    wb.save(str(path))


def read_bundle_id_marker(path: Path) -> str | None:
    """Return embedded bundle_id from workbook at *path*, or None if missing."""
    wb = openpyxl.load_workbook(str(path), data_only=False, read_only=True)
    if XLCLOAK_META_SHEET not in wb.sheetnames:
        return None
    ws = wb[XLCLOAK_META_SHEET]
    if ws[XLCLOAK_META_KEY_CELL].value != XLCLOAK_BUNDLE_ID_KEY:
        return None
    value = ws[XLCLOAK_META_VALUE_CELL].value
    if value is None:
        return None
    return str(value)
